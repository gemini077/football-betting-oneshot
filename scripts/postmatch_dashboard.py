#!/usr/bin/env python3
"""Convert the latest post-match Excel workbook into a self-contained HTML dashboard."""

from __future__ import annotations

import argparse
import html
import json
import re
import shutil
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
try:
    from paper_ledger import pair_key
except ImportError:
    from scripts.paper_ledger import pair_key


BASE_DIR = Path(__file__).resolve().parents[1]
RUNTIME_PATH = BASE_DIR / "05_RUNTIME_STATE.json"
QUEUE_PATH = BASE_DIR / "data" / "postmatch_automation" / "queue.json"
OUTPUT_ROOT = BASE_DIR / "data" / "postmatch_dashboard"
REPORT_ROOT = BASE_DIR / "data" / "postmatch_reports"
SHANGHAI = timezone(timedelta(hours=8))


def load_json(path: Path, default: Any = None) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return default


def latest_workbook(runtime: dict[str, Any], explicit: Path | None) -> Path:
    if explicit:
        return explicit if explicit.is_absolute() else BASE_DIR / explicit
    configured = runtime.get("latest_review_workbook")
    if configured:
        candidate = BASE_DIR / configured
        if candidate.exists():
            return candidate
    candidates = sorted((BASE_DIR / "data" / "postmatch_reviews").glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError("No post-match workbook found")
    return candidates[0]


def table_rows(ws, header_row: int, max_rows: int = 200) -> list[dict[str, Any]]:
    headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_index in range(header_row + 1, min(ws.max_row, header_row + max_rows) + 1):
        first = ws.cell(row_index, 1).value
        if first in (None, ""):
            continue
        row: dict[str, Any] = {}
        for col, header_value in enumerate(headers, 1):
            key = str(header_value) if header_value not in (None, "") else f"col_{col}"
            value = ws.cell(row_index, col).value
            if isinstance(value, datetime):
                value = value.isoformat()
            row[key] = value
        rows.append(row)
    return rows


def text(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, float):
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def esc(value: Any) -> str:
    return html.escape(text(value))


def money(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def calculate_kpis(locks: list[dict[str, Any]], signals: list[dict[str, Any]], runtime: dict[str, Any]) -> dict[str, Any]:
    settled = [row for row in locks if text(row.get("注单状态")) in {"红", "黑", "走水", "半红", "半黑"}]
    stake = sum(money(row.get("下注金额")) for row in settled)
    returned = sum(money(row.get("实际回收金额")) for row in settled)
    hits = sum(1 if row.get("注单状态") == "红" else 0.5 if row.get("注单状态") == "半红" else 0 for row in settled)
    bankroll = runtime.get("bankroll", {})
    return {
        "reviews": len(signals),
        "locked": len(settled),
        "hit_rate": (hits / len(settled) * 100) if settled else None,
        "stake": stake,
        "returned": returned,
        "profit": returned - stake,
        "roi": ((returned - stake) / stake * 100) if stake else None,
        "balance": money(bankroll.get("current_balance")),
    }


def badge_class(value: Any) -> str:
    value = text(value)
    if "不可计入" in value or "不可核验" in value:
        return "neutral"
    if value.startswith(("半红", "半黑", "走水")) or value in {"观察", "局部正确"}:
        return "warn"
    if value.startswith("否") or value in {"黑", "逻辑错误"} or "未中" in value:
        return "bad"
    if value.startswith("是") or value in {"红", "逻辑正确", "已启用"} or "命中" in value:
        return "good"
    return "neutral"


def safe_slug(value: Any) -> str:
    return re.sub(r"[^0-9A-Za-z._-]+", "_", str(value or "review")).strip("_") or "review"


def row_match_id(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        match = re.search(r"(?:^|[^A-Za-z0-9])(FBOS-\d{12}-[0-9a-f]{10}|M\d+|shuju[_:]?\d+|fixture[_:]?[^｜|\s]+)", text(row.get(key)), re.IGNORECASE)
        if match:
            return match.group(1).replace(":", "_")
    return ""


def review_data(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, data_only=True, read_only=False)
    locks = table_rows(wb["01_锁单与结算明细"], 3)
    signals = table_rows(wb["02_赛前信号与赛果归因"], 3)
    timeline = table_rows(wb["03_时间轴与盘口验证"], 3)
    roots = table_rows(wb["05_根因决策树与修正池"], 3)

    known = {str(row.get("MatchID") or "") for row in signals}
    for path in sorted((BASE_DIR / "data" / "postmatch_reviews").glob("*.json")):
        payload = load_json(path, {}) or {}
        match_id = str(payload.get("MatchID") or payload.get("match_key") or path.stem)
        if not payload.get("赛事与对阵") or match_id in known:
            continue
        payload["MatchID"] = match_id
        payload["kickoff_local"] = (payload.get("match") or {}).get("kickoff_local")
        signals.append(payload)
        if isinstance(payload.get("_timeline"), dict):
            timeline.append({"记录ID": match_id, "比赛ID与对阵": f"{match_id}｜{payload.get('赛事与对阵')}", **payload["_timeline"]})
        if isinstance(payload.get("_root_cause"), dict):
            roots.append({"比赛场次": f"{match_id}｜{payload.get('赛事与对阵')}", **payload["_root_cause"]})
        known.add(match_id)
    runtime = load_json(RUNTIME_PATH, {}) or {}
    reviewed = runtime.get("latest_reviewed_matches") or []
    for row in signals:
        if row.get("kickoff_local"): continue
        name = text(row.get("赛事与对阵"))
        match = next((item for item in reviewed if text(item.get("match")) == name), None)
        if match: row["kickoff_local"] = match.get("kickoff")
    ordered = sorted(signals, key=lambda row: text(row.get("kickoff_local")) if row.get("kickoff_local") else "", reverse=True)
    for index, row in enumerate(ordered, 1):
        kickoff = text(row.get("kickoff_local"))
        stamp = re.sub(r"\D", "", kickoff)[:8] if kickoff != "—" else "HISTORY"
        row["display_id"] = f"{stamp}-{index:03d}"
    return locks, ordered, timeline, roots


def _field(label: str, value: Any, *, accent: bool = False) -> str:
    return f'<div class="metric{" accent" if accent else ""}"><small>{esc(label)}</small><strong>{esc(value)}</strong></div>'


def settlement_label(value: Any) -> str:
    if value is True:
        return "命中"
    if value is False:
        return "未命中"
    if value is None:
        return "不可结算"
    return text(value)


def render_review_page(signal: dict[str, Any], timeline: dict[str, Any], root: dict[str, Any], locks: list[dict[str, Any]], generated_at: datetime, paper_tickets: list[dict] | None = None, real_bets: list[dict] | None = None) -> str:
    match_id = str(signal.get("display_id") or "—")
    match_name = signal.get("赛事与对阵") or "未命名比赛"
    score = signal.get("实际90分钟比分")
    classification = signal.get("红黑与模型逻辑分类")
    settlement = signal.get("settlement") if isinstance(signal.get("settlement"), dict) else {}

    audit_rows = [
        ("唯一主维度", signal.get("赛前首推主维度"), signal.get("主维度是否命中")),
        ("唯一正确比分", signal.get("赛前唯一首推比分"), signal.get("比分是否命中")),
        ("亚洲让球", signal.get("赛前亚盘方向"), (settlement.get("asian_handicap") or {}).get("hit") if settlement else "按复盘摘要核验"),
        ("大小球/总进球", signal.get("赛前大小球方向"), (settlement.get("total_goals_mode") or {}).get("hit") if settlement else "按复盘摘要核验"),
        ("双方进球", signal.get("赛前BTTS判断"), (settlement.get("btts") or {}).get("hit") if settlement else "按复盘摘要核验"),
    ]
    audit_html = "".join(
        f'<tr><td>{esc(market)}</td><td>{esc(pick)}</td><td><span class="pill {badge_class(settlement_label(result))}">{esc(settlement_label(result))}</span></td></tr>'
        for market, pick, result in audit_rows
    )
    timeline_fields = [
        ("快照覆盖", timeline.get("快照覆盖")),
        ("判断如何变化", timeline.get("判断如何变化")),
        ("临盘资金与机构行为", timeline.get("临盘资金与机构行为")),
        ("对最终判断的影响", timeline.get("对最终判断的影响")),
        ("最后有效判断", timeline.get("最后有效判断")),
        ("数据有效性", timeline.get("数据有效性")),
        ("初盘定位", timeline.get("初盘定位")),
        ("最后赛前快照", timeline.get("终盘定位（最后赛前快照）") or timeline.get("终盘定位（临场15min）")),
        ("完整变化轨迹", timeline.get("终盘对比初盘变化")),
        ("赛果验证", timeline.get("最终赛果验证")),
        ("来源", timeline.get("来源说明") or timeline.get("来源/备注")),
    ]
    timeline_html = "".join(
        f'<div class="fact"><b>{esc(label)}</b><span>{esc(value)}</span></div>'
        for label, value in timeline_fields if value not in (None, "", "—")
    ) or '<div class="empty-cell">没有可复核的赛前节点记录</div>'
    root_fields = [
        ("结算错项", root.get("结算错项")),
        ("最可能根因", root.get("最可能根因")),
        ("赛前已知风险", root.get("赛前已知风险")),
        ("反事实条件", root.get("反事实条件")),
        ("比分误差定位", root.get("比分误差定位")),
        ("模型修正", root.get("模型修正")),
        ("修正状态", root.get("修正状态")),
        ("复盘结论", root.get("复盘结论")),
        ("决策节点审计", root.get("决策节点审计")),
        ("反事实推演", root.get("反事实推演")),
        ("具体修改建议", root.get("具体修改建议")),
        ("最大错点类型", root.get("最大错点类型") or signal.get("最大错点类型")),
    ]
    root_html = "".join(
        f'<div class="fact"><b>{esc(label)}</b><span>{esc(value)}</span></div>'
        for label, value in root_fields if value not in (None, "", "—")
    ) or '<div class="empty-cell">本场尚未形成可验证的根因结论</div>'
    lock_rows = "".join(
        f'<tr><td>{esc(row.get("注单ID"))}</td><td>{esc(row.get("投注层标签"))}</td><td>{esc(row.get("投注方向"))}</td><td>{esc(row.get("下注赔率"))}</td><td>{esc(row.get("下注金额"))}</td><td>{esc(row.get("注单状态"))}</td><td>{esc(row.get("盈亏"))}</td></tr>'
        for row in locks
    ) or '<tr><td colspan="7" class="empty-cell">本场未锁单；只复盘模型，不制造真实盈亏。</td></tr>'
    paper_rows = "".join(
        f'<tr><td>{esc(row.get("ticket_id"))}</td><td>{esc(row.get("market"))}</td><td>{esc(row.get("selection"))}</td><td>{esc(row.get("odds"))}</td><td>{esc(row.get("stake_units"))}</td><td>{esc(row.get("status"))}</td><td>{esc(row.get("profit_units"))}</td></tr>'
        for row in (paper_tickets or [])
    ) or '<tr><td colspan="7" class="empty-cell">本场没有通过正EV与最低金额审核的T-90模拟注单。</td></tr>'
    real_rows = "".join(
        f'<tr><td>{esc(row.get("bet_id"))}</td><td>{esc(row.get("market"))}</td><td>{esc(row.get("selection"))}</td><td>{esc(row.get("odds"))}</td><td>{esc(row.get("stake"))}</td><td>{esc(row.get("status"))}</td><td>{esc(row.get("profit"))}</td></tr>'
        for row in (real_bets or [])
    ) or lock_rows
    return f'''<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(match_name)}｜完整赛后复盘</title><style>
:root{{--bg:#0a0811;--panel:#151020;--panel2:#1d142b;--ink:#f8f3fb;--body:#cdc4d7;--mut:#948aa2;--line:rgba(255,255,255,.1);--red:#ff3657;--pink:#ff7189;--purple:#9b7fd0;--green:#39d6a0;--amber:#ffbd5c}}
*{{box-sizing:border-box}}body{{margin:0;padding:30px;background:radial-gradient(70% 40% at 50% -5%,rgba(255,54,87,.24),transparent 70%),linear-gradient(180deg,#181025,#0a0811 52%);color:var(--body);font:14px/1.65 "Segoe UI","Microsoft YaHei",sans-serif}}.page{{max-width:1380px;margin:auto}}a{{color:#d9c8f2;text-decoration:none}}.back{{display:inline-flex;padding:8px 12px;border:1px solid rgba(155,127,208,.45);border-radius:9px;background:rgba(155,127,208,.08);margin-bottom:16px}}.hero{{position:relative;overflow:hidden;border:1px solid var(--line);border-radius:22px;padding:42px;background:linear-gradient(145deg,rgba(255,54,87,.18),transparent 42%),linear-gradient(25deg,#100b1b,#211232)}}.eyebrow{{color:var(--pink);font-size:12px;font-weight:800;letter-spacing:.25em}}h1{{color:var(--ink);font-size:clamp(30px,5vw,62px);line-height:1.08;margin:14px 0 8px}}.score{{font-size:clamp(38px,7vw,78px);font-weight:900;color:var(--red);line-height:1}}.meta{{color:var(--mut)}}.grid{{display:grid;grid-template-columns:repeat(12,minmax(0,1fr));gap:18px;margin-top:20px}}.card{{grid-column:span 6;padding:22px;border:1px solid var(--line);border-radius:16px;background:linear-gradient(180deg,rgba(255,255,255,.035),rgba(255,255,255,.018));overflow:hidden}}.full{{grid-column:span 12}}h2{{color:var(--ink);font-size:17px;margin:0 0 16px}}.metrics{{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-top:22px}}.metric{{padding:14px;border:1px solid var(--line);border-radius:11px;background:rgba(255,255,255,.025)}}.metric.accent{{border-color:rgba(255,54,87,.45);background:rgba(255,54,87,.07)}}.metric small{{display:block;color:var(--mut);font-size:11px}}.metric strong{{display:block;color:var(--ink);font-size:15px;margin-top:5px}}.summary{{padding:18px;border-left:3px solid var(--red);background:rgba(255,54,87,.06);border-radius:0 12px 12px 0;color:var(--ink);font-size:16px}}.table-wrap{{overflow:auto;border:1px solid var(--line);border-radius:11px}}table{{width:100%;border-collapse:collapse;min-width:720px}}th,td{{padding:11px 12px;border-bottom:1px solid var(--line);text-align:left}}th{{color:var(--mut);font-size:11px;background:rgba(255,255,255,.035)}}.pill{{display:inline-flex;padding:4px 9px;border-radius:999px;background:#273047}}.pill.good{{color:var(--green);background:rgba(57,214,160,.1)}}.pill.bad{{color:#ff8e9f;background:rgba(255,54,87,.12)}}.pill.warn{{color:var(--amber);background:rgba(255,189,92,.1)}}.facts{{display:grid;gap:0}}.fact{{display:grid;grid-template-columns:150px 1fr;gap:18px;padding:12px;border-bottom:1px solid var(--line)}}.fact b{{color:var(--mut)}}.fact span{{color:var(--ink)}}.empty-cell{{text-align:center;color:var(--mut)}}footer{{margin-top:24px;padding-top:16px;border-top:1px solid var(--line);color:var(--mut);font-size:12px}}
@media(max-width:900px){{body{{padding:15px}}.hero{{padding:28px 22px}}.card,.full{{grid-column:span 12}}.metrics{{grid-template-columns:repeat(2,1fr)}}}}@media(max-width:540px){{.metrics{{grid-template-columns:1fr}}.fact{{grid-template-columns:1fr;gap:4px}}}}
</style></head><body><main class="page"><a class="back" href="../match_workspace/latest.html">← 返回比赛工作台</a>
<header class="hero"><div class="eyebrow">FOOTBALL BETTING ONESHOT · POST-MATCH REVIEW</div><h1>{esc(match_name)}</h1><div class="score">{esc(score)}</div><div class="meta">{esc(match_id)} · 90分钟含伤停，不含加时与点球 · 生成于 {esc(generated_at.strftime('%Y-%m-%d %H:%M:%S'))}</div><div class="metrics">{_field('主维度结算',signal.get('主维度是否命中'),accent=True)}{_field('唯一比分',signal.get('比分是否命中'))}{_field('模型逻辑',classification)}{_field('最大错点',signal.get('最大错点类型'))}{_field('随机事件',signal.get('随机事件剔除'))}{_field('样本有效性',signal.get('是否有效样本'))}</div></header>
<div class="grid"><section class="card full"><h2>01｜复盘结论</h2><div class="summary">{esc(signal.get('复盘摘要'))}</div></section>
<section class="card full"><h2>02｜全部玩法严格结算</h2><div class="table-wrap"><table><thead><tr><th>玩法</th><th>赛前冻结判断</th><th>赛后严格结算</th></tr></thead><tbody>{audit_html}</tbody></table></div><p class="meta">相邻比分只用于误差诊断，永远不算精确命中；辅助玩法命中不替代唯一主维度。</p></section>
<section class="card"><h2>03｜赛前推理回放</h2><div class="facts">{_field('三维交叉验证',signal.get('赛前三维交叉验证结论'))}{_field('盘路性质',signal.get('盘路性质判定'))}{_field('赛前最大错点',signal.get('赛前最大错点'))}{_field('赛后错点归因',signal.get('错点归因（单选）'))}{_field('相邻比分污染',signal.get('相邻比分污染'))}{_field('冷门/右尾污染',signal.get('冷门/右尾污染'))}</div></section>
<section class="card"><h2>04｜盘口时间线与数据有效性</h2><div class="facts">{timeline_html}</div></section>
<section class="card full"><h2>05｜根因、反事实与模型修正</h2><div class="facts">{root_html}</div></section>
<section class="card full"><h2>06｜模拟注单结算</h2><div class="table-wrap"><table><thead><tr><th>模拟ID</th><th>玩法</th><th>方向</th><th>冻结赔率</th><th>金额</th><th>状态</th><th>盈亏</th></tr></thead><tbody>{paper_rows}</tbody></table></div></section>
<section class="card full"><h2>07｜真实注单结算</h2><div class="table-wrap"><table><thead><tr><th>注单ID</th><th>玩法</th><th>方向</th><th>实际赔率</th><th>实际金额</th><th>状态</th><th>盈亏</th></tr></thead><tbody>{real_rows}</tbody></table></div></section></div>
<footer>冻结赛前判断后再核验赛果；没有的价格和时间节点保持为空，不用赛后信息回填。未明确“锁单/已下单”时，不生成真实注单或账户盈亏。</footer></main></body></html>'''


def write_review_pages(signals: list[dict[str, Any]], timelines: list[dict[str, Any]], roots: list[dict[str, Any]], locks: list[dict[str, Any]], generated_at: datetime, output_root: Path = REPORT_ROOT) -> dict[str, str]:
    output_root.mkdir(parents=True, exist_ok=True)
    timeline_map = {row_match_id(row, "比赛ID与对阵", "记录ID"): row for row in timelines}
    root_map = {row_match_id(row, "比赛场次"): row for row in roots}
    links: dict[str, str] = {}
    paper = (load_json(BASE_DIR / "data" / "paper_ledger" / "latest.json", {}) or {}).get("tickets") or []
    real = (load_json(BASE_DIR / "data" / "real_bets" / "latest.json", {}) or {}).get("bets") or []
    for signal in signals:
        match_id = str(signal.get("MatchID") or safe_slug(signal.get("赛事与对阵")))
        related_locks = [row for row in locks if match_id and match_id in text(row.get("比赛ID与对阵"))]
        name = text(signal.get("赛事与对阵"))
        teams = re.split(r"\s+vs\s+", name, maxsplit=1, flags=re.IGNORECASE)
        key = pair_key(*teams) if len(teams) == 2 else ""
        related_paper = [row for row in paper if row.get("match_key") == key]
        related_real = [row for row in real if text(row.get("match")) == name]
        target = output_root / f"{safe_slug(match_id)}.html"
        target.write_text(render_review_page(signal, timeline_map.get(match_id, {}), root_map.get(match_id, {}), related_locks, generated_at, related_paper, related_real), encoding="utf-8")
        links[match_id] = f"../postmatch_reports/{target.name}"
    return links


def signal_cards(signals: list[dict[str, Any]], report_links: dict[str, str] | None = None) -> str:
    report_links = report_links or {}
    cards = []
    for row in signals:
        search = " ".join(text(v) for v in row.values()).casefold()
        report_url = report_links.get(str(row.get("MatchID") or ""), "")
        cards.append(
            f'''<article class="match-card" data-search="{html.escape(search, quote=True)}" data-hit="{esc(row.get('主维度是否命中'))}">
              <div class="card-head"><span class="match-id">{esc(row.get('display_id'))}</span><h3>{esc(row.get('赛事与对阵'))}</h3><span class="badge {badge_class(row.get('主维度是否命中'))}">主维度严格结算 {esc(row.get('主维度是否命中'))}</span></div>
              <div class="meta">开赛：{esc(row.get('kickoff_local'))}</div>
              <div class="score-strip"><span>赛前唯一比分 <b>{esc(row.get('赛前唯一首推比分'))}</b></span><span>实际90分钟 <b>{esc(row.get('实际90分钟比分'))}</b></span><span class="badge {badge_class(row.get('比分是否命中'))}">唯一比分精确命中 {esc(row.get('比分是否命中'))}</span></div>
              <div class="detail-grid">
                <div><label>赛前主维度</label><p>{esc(row.get('赛前首推主维度'))}</p></div>
                <div><label>亚盘方向</label><p>{esc(row.get('赛前亚盘方向'))}</p></div>
                <div><label>大小球</label><p>{esc(row.get('赛前大小球方向'))}</p></div>
                <div><label>BTTS</label><p>{esc(row.get('赛前BTTS判断'))}</p></div>
              </div>
              <p class="summary">{esc(row.get('复盘摘要'))}</p>
              <div class="card-foot"><span class="badge {badge_class(row.get('红黑与模型逻辑分类'))}">{esc(row.get('红黑与模型逻辑分类'))}</span><span>最大错点：{esc(row.get('最大错点类型'))}</span>{f'<a class="report-link" href="{html.escape(report_url, quote=True)}">打开完整复盘 →</a>' if report_url else ''}</div>
            </article>'''
        )
    return "\n".join(cards)


def data_table(rows: list[dict[str, Any]], columns: list[str], table_id: str) -> str:
    headers = "".join(f"<th>{esc(column)}</th>" for column in columns)
    body = []
    for row in rows:
        search = " ".join(text(row.get(column)) for column in columns).casefold()
        cells = "".join(f"<td>{esc(row.get(column))}</td>" for column in columns)
        body.append(f'<tr data-search="{html.escape(search, quote=True)}">{cells}</tr>')
    return f'<div class="table-wrap"><table id="{table_id}"><thead><tr>{headers}</tr></thead><tbody>{"".join(body)}</tbody></table></div>'


def render_html(workbook_path: Path, runtime: dict[str, Any], queue: dict[str, Any], generated_at: datetime,
                review_rows: tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]] | None = None,
                report_links: dict[str, str] | None = None) -> str:
    locks, signals, timeline, roots = review_rows or review_data(workbook_path)
    kpis = calculate_kpis(locks, signals, runtime)
    display_map = {str(row.get("MatchID") or ""): str(row.get("display_id") or "") for row in signals}
    def public_value(value: Any) -> Any:
        rendered = text(value)
        for internal, public in display_map.items():
            if internal and public:
                rendered = rendered.replace(internal, public)
        return rendered
    timeline_display = [{key: public_value(value) for key, value in row.items()} for row in timeline]
    roots_display = [{key: public_value(value) for key, value in row.items()} for row in roots]
    pending = len(queue.get("pending", []))
    waiting = queue.get("counts", {}).get("waiting_for_finish", 0)
    hit_rate = "—" if kpis["hit_rate"] is None else f'{kpis["hit_rate"]:.1f}%'
    roi = "—" if kpis["roi"] is None else f'{kpis["roi"]:.1f}%'
    lock_table = data_table(locks, ["注单ID", "比赛ID与对阵", "主维度玩法", "投注方向", "下注赔率", "下注金额", "赛果", "注单状态", "实际回收金额", "模型逻辑分类", "备注"], "locks-table")
    timeline_table = data_table(timeline_display, ["记录ID", "比赛ID与对阵", "开赛倒计时", "锁单窗口合规性", "初盘定位", "终盘定位（临场15min）", "终盘对比初盘变化", "最终赛果验证", "数据完整度"], "timeline-table")
    root_table = data_table(roots_display, ["比赛场次", "决策节点审计", "反事实推演", "赛前可识别性", "是否修改模型", "具体修改建议", "收敛结论", "最大错点类型", "生效状态", "优先级"], "root-table")
    source_name = esc(workbook_path.name)
    updated = generated_at.strftime("%Y-%m-%d %H:%M:%S")
    model_version = esc(runtime.get("model_version"))
    return f'''<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Football Betting OneShot｜赛后复盘工作台</title>
<style>
:root{{--bg:#090d18;--panel:#11182a;--panel2:#161f35;--line:#293653;--text:#edf2ff;--muted:#94a3be;--accent:#7c3aed;--accent2:#ec4899;--good:#21c77a;--bad:#fb5b68;--warn:#f4b942;--blue:#46a5ff}}
*{{box-sizing:border-box}} body{{margin:0;background:radial-gradient(circle at 12% 0,#22153f 0,transparent 32%),radial-gradient(circle at 90% 10%,#182f52 0,transparent 28%),var(--bg);color:var(--text);font:14px/1.55 "Microsoft YaHei",system-ui,sans-serif}}
    .shell{{max-width:1500px;margin:auto;padding:28px}} .hero{{display:flex;justify-content:space-between;gap:24px;align-items:flex-end;margin-bottom:22px}} .eyebrow{{color:#c4b5fd;font-weight:700;letter-spacing:.12em}} h1{{font-size:34px;margin:6px 0}} .hero p{{margin:0;color:var(--muted)}} .sync{{min-width:280px;background:#0e1526;border:1px solid var(--line);border-radius:16px;padding:15px 18px}} .sync strong{{display:block;color:var(--good);font-size:16px}} .kpis{{display:grid;grid-template-columns:repeat(7,minmax(120px,1fr));gap:12px;margin-bottom:20px}} .kpi{{background:linear-gradient(145deg,var(--panel2),var(--panel));border:1px solid var(--line);border-radius:16px;padding:16px}} .kpi span{{color:var(--muted);font-size:12px}} .kpi b{{display:block;font-size:23px;margin-top:5px}} .toolbar{{position:sticky;top:0;z-index:9;background:rgba(9,13,24,.88);backdrop-filter:blur(12px);display:flex;gap:10px;padding:12px 0}} button,.filter,input{{border:1px solid var(--line);background:#11192b;color:var(--text);border-radius:10px;padding:10px 13px}} button{{cursor:pointer}} button.active{{background:linear-gradient(90deg,var(--accent),var(--accent2));border-color:transparent}} input{{flex:1;min-width:220px}} .view{{display:none}} .view.active{{display:block}} .match-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px}} .match-card{{background:linear-gradient(150deg,#151e34,#0f1626);border:1px solid var(--line);border-radius:18px;padding:18px;box-shadow:0 14px 40px rgba(0,0,0,.16)}} .card-head{{display:flex;align-items:center;gap:10px}} .card-head h3{{font-size:17px;margin:0;flex:1}} .match-id{{font-weight:800;color:#c4b5fd}} .score-strip{{display:flex;gap:18px;background:#0b1120;border-radius:12px;padding:12px;margin:14px 0}} .score-strip b{{font-size:17px}} .detail-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:8px}} .detail-grid div{{border-left:2px solid #3b4b6c;padding-left:9px}} label{{color:var(--muted);font-size:11px}} p{{margin:3px 0}} .summary{{color:#c8d1e5;margin:15px 0}} .card-foot{{display:flex;align-items:center;justify-content:space-between;gap:10px;color:var(--muted);flex-wrap:wrap}} .report-link{{margin-left:auto;color:#fff;text-decoration:none;border:1px solid #7c3aed;background:linear-gradient(90deg,#7c3aed,#ec4899);border-radius:10px;padding:8px 12px;font-weight:700}} .report-link:hover{{filter:brightness(1.12)}} .badge{{display:inline-flex;align-items:center;padding:3px 9px;border-radius:999px;font-size:12px;white-space:nowrap}} .badge.good{{background:#123b2c;color:#62e4a6}} .badge.bad{{background:#451d27;color:#ff8c98}} .badge.warn{{background:#453717;color:#ffd36e}} .badge.neutral{{background:#25324a;color:#c1cbe0}} .table-wrap{{overflow:auto;background:#0f1626;border:1px solid var(--line);border-radius:16px}} table{{border-collapse:collapse;width:100%;min-width:1100px}} th{{position:sticky;top:0;background:#1b2540;color:#cbd5e9;text-align:left;padding:12px;border-bottom:1px solid var(--line);font-size:12px}} td{{padding:11px 12px;border-bottom:1px solid #222e48;vertical-align:top;max-width:320px}} tr:hover td{{background:#141e33}} .section-head{{display:flex;justify-content:space-between;align-items:center;margin:10px 0 14px}} .section-head h2{{margin:0}} .empty{{display:none;padding:35px;text-align:center;color:var(--muted)}} footer{{color:var(--muted);margin-top:22px;padding:15px 0;border-top:1px solid var(--line)}}
@media(max-width:1050px){{.kpis{{grid-template-columns:repeat(3,1fr)}}.match-grid{{grid-template-columns:1fr}}.detail-grid{{grid-template-columns:repeat(2,1fr)}}}} @media(max-width:650px){{.shell{{padding:16px}}.hero{{display:block}}.sync{{margin-top:14px}}.kpis{{grid-template-columns:repeat(2,1fr)}}.toolbar{{flex-wrap:wrap}}.score-strip{{flex-wrap:wrap}}}}
</style></head><body><main class="shell">
<header class="hero"><div><div class="eyebrow">FOOTBALL BETTING ONESHOT · {model_version}</div><h1>赛后复盘工作台</h1><p>按开赛时间由近到远｜90分钟赛果口径｜赛前冻结、赛后归因</p></div><div class="sync"><strong>● 自动复盘已接入</strong><span>待核验 {pending} 场 · 等待结束 {waiting} 场</span><br><small>页面生成：{updated}</small></div></header>
<section class="kpis"><div class="kpi"><span>有效复盘</span><b>{kpis['reviews']}</b></div><div class="kpi"><span>已结注单</span><b>{kpis['locked']}</b></div><div class="kpi"><span>注单命中率</span><b>{hit_rate}</b></div><div class="kpi"><span>累计投注</span><b>¥{kpis['stake']:.2f}</b></div><div class="kpi"><span>净盈亏</span><b>¥{kpis['profit']:.2f}</b></div><div class="kpi"><span>ROI</span><b>{roi}</b></div><div class="kpi"><span>当前余额</span><b>¥{kpis['balance']:.2f}</b></div></section>
<nav class="toolbar"><button class="tab active" data-view="signals">比赛复盘</button><button class="tab" data-view="locks">注单结算</button><button class="tab" data-view="timeline">盘口时间轴</button><button class="tab" data-view="roots">根因与修正</button><input id="search" placeholder="搜索球队、记录ID、玩法或错点…"><select id="hit-filter" class="filter"><option value="">全部命中状态</option><option>是</option><option>否</option></select></nav>
<section id="signals" class="view active"><div class="section-head"><h2>赛前信号与赛果归因</h2><span>{len(signals)} 场</span></div><div class="match-grid">{signal_cards(signals, report_links)}</div></section>
<section id="locks" class="view"><div class="section-head"><h2>锁单与结算明细</h2><span>{len(locks)} 单</span></div>{lock_table}</section>
<section id="timeline" class="view"><div class="section-head"><h2>时间轴与盘口有效性</h2><span>{len(timeline)} 条</span></div>{timeline_table}</section>
<section id="roots" class="view"><div class="section-head"><h2>根因决策树与修正池</h2><span>{len(roots)} 条</span></div>{root_table}</section>
<div id="empty" class="empty">没有符合筛选条件的记录</div><footer>数据源：{source_name}｜本页为Excel的只读可视化，不改变锁单、账户余额或执行状态。</footer>
</main><script>
const tabs=[...document.querySelectorAll('.tab')], views=[...document.querySelectorAll('.view')], search=document.querySelector('#search'), hit=document.querySelector('#hit-filter'), empty=document.querySelector('#empty');
tabs.forEach(b=>b.onclick=()=>{{tabs.forEach(x=>x.classList.toggle('active',x===b));views.forEach(v=>v.classList.toggle('active',v.id===b.dataset.view));applyFilter();}});
function applyFilter(){{const q=search.value.trim().toLowerCase(), active=document.querySelector('.view.active');let shown=0;active.querySelectorAll('[data-search]').forEach(el=>{{const ok=el.dataset.search.includes(q)&&(!hit.value||active.id!=='signals'||el.dataset.hit===hit.value);el.style.display=ok?'':'none';if(ok)shown++;}});empty.style.display=shown?'none':'block';}}
search.addEventListener('input',applyFilter);hit.addEventListener('change',applyFilter);
</script></body></html>'''


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--output-root", type=Path, default=OUTPUT_ROOT)
    args = parser.parse_args()
    runtime = load_json(RUNTIME_PATH, {})
    queue = load_json(QUEUE_PATH, {})
    workbook = latest_workbook(runtime, args.workbook)
    if not workbook.exists():
        raise SystemExit(f"Workbook not found: {workbook}")
    now = datetime.now(SHANGHAI)
    run_id = now.strftime("%Y%m%d_%H%M%S")
    output_root = args.output_root if args.output_root.is_absolute() else BASE_DIR / args.output_root
    snapshot_dir = output_root / run_id
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    rows = review_data(workbook)
    report_links = write_review_pages(rows[1], rows[2], rows[3], rows[0], now)
    content = render_html(workbook, runtime, queue, now, rows, report_links)
    snapshot = snapshot_dir / "index.html"
    snapshot.write_text(content, encoding="utf-8")
    latest = output_root / "latest.html"
    shutil.copyfile(snapshot, latest)
    metadata = {
        "schema_version": "1.0",
        "generated_at": now.isoformat(),
        "source_workbook": workbook.relative_to(BASE_DIR).as_posix(),
        "snapshot": snapshot.relative_to(BASE_DIR).as_posix(),
        "latest": latest.relative_to(BASE_DIR).as_posix(),
        "individual_reports": len(report_links),
        "individual_report_root": REPORT_ROOT.relative_to(BASE_DIR).as_posix(),
    }
    (snapshot_dir / "manifest.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(metadata, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
