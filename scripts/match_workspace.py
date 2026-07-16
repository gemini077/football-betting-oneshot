#!/usr/bin/env python3
"""Build one read-only workspace for Sporttery schedule, pre-match reports and reviews."""

from __future__ import annotations

import argparse
import html
import json
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

try:
    from paper_ledger import build_paper_ledger, pair_key, parse_score
    from paper_channel_prices import sync_channel_price_overrides
except ImportError:  # package import used by tests
    from scripts.paper_ledger import build_paper_ledger, pair_key, parse_score
    from scripts.paper_channel_prices import sync_channel_price_overrides

try:
    from openpyxl import load_workbook
except ImportError:  # bundled desktop runtime may not be on a plain Python PATH
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUTPUT = DATA / "match_workspace"
RUNTIME = ROOT / "05_RUNTIME_STATE.json"
SHANGHAI = ZoneInfo("Asia/Shanghai")


def parse_kickoff_local(value: Any) -> datetime | None:
    text = str(value or "").strip().replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).replace(tzinfo=SHANGHAI)
        except ValueError:
            continue
    return None


def match_should_be_finished(kickoff: Any, now: datetime) -> bool:
    parsed = parse_kickoff_local(kickoff)
    return parsed is not None and now >= parsed + timedelta(hours=2, minutes=15)

PORTFOLIO_LAYERS = {
    "保本层": {
        "display": "保本层｜低波动",
        "color": "#39d6a0",
        "role": "当天组合的稳定器，不代表保证本金。",
        "purpose": "当天组合的稳定器，不代表保证本金。",
        "guardrail": "不得用多场热门串关伪装低风险。",
    },
    "中轴层": {
        "display": "中轴层｜主收益",
        "color": "#ffbd5c",
        "role": "承载当天主要正EV暴露，可为单关或合格串关。",
        "purpose": "承载当天主要正EV暴露，可为单关或合格串关。",
        "guardrail": "每一腿必须先独立通过价格审核。",
    },
    "博上层": {
        "display": "博上层｜高方差",
        "color": "#ff7189",
        "role": "小额高赔率正EV；正确比分只能进入本层。",
        "purpose": "小额高赔率正EV；正确比分只能进入本层。",
        "guardrail": "不得承担回本任务或覆盖一串比分。",
    },
}


def load_json(path: Path, default: Any = None) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return default


def norm(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]", "", str(value or "").casefold())


def esc(value: Any) -> str:
    return html.escape(str(value if value not in (None, "") else "—"))


def latest_schedule(target_date: str) -> tuple[Path | None, dict]:
    candidates = list((DATA / "schedule_updates").glob(f"**/*{target_date}*.json"))
    candidates += list((DATA / "fetch_runs").glob(f"**/*sporttery_{target_date}.json"))
    rows: list[tuple[int, float, Path, dict]] = []
    for path in candidates:
        payload = load_json(path, {})
        if payload.get("source") != "sporttery.cn" or not payload.get("success"):
            continue
        rows.append((len(payload.get("matches") or []), path.stat().st_mtime, path, payload))
    if not rows:
        return None, {"matches": [], "date": target_date, "success": False}
    _, _, path, payload = max(rows, key=lambda row: (row[0], row[1]))
    return path, payload


def latest_reports() -> dict[str, dict]:
    reports: dict[str, dict] = {}
    for path in sorted((DATA / "analysis_reports").glob("*/*.json"), key=lambda p: p.stat().st_mtime):
        payload = load_json(path, {})
        match = payload.get("match") or {}
        if not match.get("home") or not match.get("away"):
            continue
        key = f"{norm(match.get('home'))}|{norm(match.get('away'))}"
        reports[key] = {
            "path": path,
            "html": next(path.parent.glob("*.html"), None),
            "payload": payload,
        }
    return reports


def verified_result_map() -> dict[str, tuple[int, int]]:
    """Read verified 90-minute scores without manufacturing missing results."""
    results: dict[str, tuple[int, int]] = {}
    for path in sorted((DATA / "postmatch_reviews").glob("*.json")):
        payload = load_json(path, {})
        match = payload.get("match") or {}
        score = parse_score((payload.get("result") or {}).get("score_90m"))
        if match.get("home") and match.get("away") and score is not None:
            results[pair_key(match.get("home"), match.get("away"))] = score
    for path in sorted((DATA / "postmatch_automation" / "results").glob("*.json")):
        payload = load_json(path, {})
        score = parse_score(payload.get("result_90m"))
        if payload.get("home") and payload.get("away") and score is not None:
            results[pair_key(payload.get("home"), payload.get("away"))] = score
    # Runtime rows are populated only after the post-match verification flow;
    # include them so older frozen paper tickets can settle even when their
    # review lives in the workbook rather than a standalone JSON file.
    runtime = load_json(RUNTIME, {}) or {}
    for row in runtime.get("latest_reviewed_matches") or []:
        parts = re.split(r"\s+vs\s+", str(row.get("match") or ""), maxsplit=1, flags=re.IGNORECASE)
        score = parse_score(row.get("result_90m"))
        if len(parts) == 2 and score is not None:
            results[pair_key(parts[0], parts[1])] = score
    return results


def create_unique_output_dir(output_root: Path, stamp: str) -> Path:
    """Create a timestamp directory without failing on same-second rebuilds."""
    candidate = output_root / stamp
    suffix = 1
    while candidate.exists():
        suffix += 1
        candidate = output_root / f"{stamp}_{suffix:02d}"
    candidate.mkdir(parents=True, exist_ok=False)
    return candidate


def _xlsx_sheet_values(path: Path, sheet_name: str) -> list[list[Any]]:
    """Read a worksheet with stdlib only when openpyxl is unavailable."""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    pkg_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in rels.findall(f"{{{pkg_ns}}}Relationship")
        }
        target = None
        for sheet in workbook.findall(f".//{{{main_ns}}}sheet"):
            if sheet.attrib.get("name") == sheet_name:
                target = rel_targets.get(sheet.attrib.get(f"{{{rel_ns}}}id"))
                break
        if not target:
            return []
        target = target.replace("\\", "/").lstrip("/")
        sheet_path = target if target.startswith("xl/") else f"xl/{target}"
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            strings = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in strings.findall(f"{{{main_ns}}}si"):
                shared.append("".join(node.text or "" for node in item.iter(f"{{{main_ns}}}t")))
        root = ET.fromstring(archive.read(sheet_path))
        rows: list[list[Any]] = []
        for row in root.findall(f".//{{{main_ns}}}row"):
            values: dict[int, Any] = {}
            for cell in row.findall(f"{{{main_ns}}}c"):
                ref = cell.attrib.get("r", "A1")
                letters = re.match(r"[A-Z]+", ref)
                col = 0
                for char in (letters.group(0) if letters else "A"):
                    col = col * 26 + ord(char) - 64
                cell_type = cell.attrib.get("t")
                value_node = cell.find(f"{{{main_ns}}}v")
                if cell_type == "inlineStr":
                    value = "".join(node.text or "" for node in cell.iter(f"{{{main_ns}}}t"))
                elif value_node is None:
                    value = None
                elif cell_type == "s":
                    index = int(value_node.text or 0)
                    value = shared[index] if 0 <= index < len(shared) else None
                elif cell_type in {"str", "b"}:
                    value = value_node.text
                else:
                    raw = value_node.text
                    try:
                        number = float(raw) if raw is not None else None
                        value = int(number) if number is not None and number.is_integer() else number
                    except (TypeError, ValueError):
                        value = raw
                values[col] = value
            width = max(values, default=0)
            rows.append([values.get(index) for index in range(1, width + 1)])
        return rows


def workbook_table_rows(path: Path, sheet_name: str, header_row: int = 3) -> list[dict]:
    values: list[list[Any]] | None = None
    if load_workbook is not None:
        wb = None
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb[sheet_name]
            values = [list(row) for row in ws.iter_rows(values_only=True)]
        except (OSError, KeyError, TypeError, ValueError):
            values = None
        finally:
            if wb is not None:
                wb.close()
    if values is None:
        try:
            values = _xlsx_sheet_values(path, sheet_name)
        except (OSError, KeyError, zipfile.BadZipFile, ET.ParseError):
            return []
    if len(values) < header_row:
        return []
    headers = values[header_row - 1]
    result = []
    for cells in values[header_row:]:
        if not cells or cells[0] in (None, ""):
            continue
        result.append({str(header): (cells[index] if index < len(cells) else None) for index, header in enumerate(headers) if header not in (None, "")})
    return result


def review_rows(runtime: dict) -> list[dict]:
    configured = runtime.get("latest_review_workbook")
    path = ROOT / configured if configured else None
    signals: list[dict] = []
    timelines: list[dict] = []
    roots: list[dict] = []
    if path and path.exists():
        signals = workbook_table_rows(path, "02_赛前信号与赛果归因")
        timelines = workbook_table_rows(path, "03_时间轴与盘口验证")
        roots = workbook_table_rows(path, "05_根因决策树与修正池")
    timeline_by_id = {str(row.get("记录ID") or "").replace("T", "M", 1): row for row in timelines}
    root_by_id = {str(row.get("比赛场次") or "").split("｜", 1)[0]: row for row in roots}
    for row in signals:
        match_id = str(row.get("MatchID") or "")
        row["_timeline"] = timeline_by_id.get(match_id)
        row["_root_cause"] = root_by_id.get(match_id)
    # GitHub 自动复盘先生成独立 JSON；工作簿是审计归档，不应成为主页显示
    # 已完赛结果的前置条件。JSON 放在末尾，find_review 反向查找时优先采用它。
    for review_path in sorted((DATA / "postmatch_reviews").glob("*.json")):
        payload = load_json(review_path, {})
        if payload.get("赛事与对阵") and payload.get("实际90分钟比分"):
            signals.append(payload)
    return signals


def find_review(home: str, away: str, rows: list[dict]) -> dict | None:
    left, right = norm(home), norm(away)
    best: tuple[float, dict] | None = None
    for row in reversed(rows):
        label = str(row.get("赛事与对阵") or "")
        pair = re.split(r"\s+vs\s+", label.split("｜")[-1], maxsplit=1, flags=re.IGNORECASE)
        row_home, row_away = (pair + [""])[:2]
        row_left, row_right = norm(row_home), norm(row_away)
        if left == row_left and (right == row_right or right in row_right or row_right in right):
            return row
        score = (SequenceMatcher(None, left, row_left).ratio() + SequenceMatcher(None, right, row_right).ratio()) / 2
        if score >= 0.72 and (best is None or score > best[0]):
            best = (score, row)
    return best[1] if best else None


def report_summary(report: dict | None) -> dict:
    if not report:
        return {
            "state": "未分析",
            "primary": "等待你选择后再分析",
            "error": "尚未分析，暂无模型错点",
            "betting": "未锁单",
        }
    payload = report["payload"]
    analysis = payload.get("analysis") or {}
    decisions = payload.get("decisions") or {}
    betting = payload.get("betting") or {}
    data_quality = payload.get("data_quality") or analysis.get("data_quality") or {}
    quality_status = str(data_quality.get("status") or "")
    serialized_errors = json.dumps(
        [analysis.get("errors") or [], data_quality, decisions.get("maximum_error_points") or []],
        ensure_ascii=False,
    ).upper()
    model = payload.get("model") or analysis.get("model") or {}
    market_only = quality_status == "仅市场基线" or (
        model.get("probabilities") is None and data_quality.get("only_official_odds_available") is True
    )
    no_data = any(marker in serialized_errors for marker in ("NO_DATA", "INSUFFICIENT_DATA"))
    if market_only or no_data:
        return {
            "state": "仅市场基线" if market_only else "数据不足",
            "primary": "仅有体彩市场基线，尚未形成模型结论" if market_only else "未形成有效分析",
            "error": "缺少球队、模型概率或多市场证据，不能给出首推",
            "betting": betting.get("state") or "空仓｜未锁单",
        }
    primary = (
        decisions.get("unique_primary_dimension")
        or
        analysis.get("primary_dimension")
        or analysis.get("main_recommendation")
        or analysis.get("final_conclusion")
        or next((row.get("market") for row in betting.get("price_audit") or [] if row.get("ev") is not None), None)
        or "已生成完整报告"
    )
    error_points = decisions.get("maximum_error_points") or []
    primary_error = error_points[0] if error_points else "临场信息可能使首推失效"
    return {
        "state": "已分析",
        "primary": primary,
        "error": primary_error,
        "betting": betting.get("state") or "空仓｜未锁单",
    }


def normalize_layer(value: Any, market: Any = None) -> str:
    aliases = {
        "保本": "保本层", "保本层": "保本层", "capital_preservation": "保本层",
        "中轴": "中轴层", "中轴层": "中轴层", "core": "中轴层",
        "博上": "博上层", "博上层": "博上层", "upside": "博上层",
    }
    raw = str(value or "").strip()
    if raw in aliases:
        return aliases[raw]
    market_text = str(market or "")
    return "博上层" if "比分" in market_text or "波胆" in market_text else "中轴层"


def report_candidates(report: dict | None, home: str, away: str) -> list[dict]:
    if not report:
        return []
    betting = (report.get("payload") or {}).get("betting") or {}
    result = []
    for index, raw in enumerate(betting.get("candidates") or [], start=1):
        item = dict(raw)
        item["ticket_id"] = str(item.get("ticket_id") or f"C{index:03d}")
        item["tier"] = normalize_layer(item.get("tier") or item.get("layer"), item.get("market"))
        item["match"] = item.get("match") or f"{home} vs {away}"
        item["status"] = item.get("status") or item.get("reprice_status") or "候选｜未锁单"
        item["form"] = item.get("form") or "单关"
        result.append(item)
    return result


def build_daily_portfolio(matches: list[dict], runtime: dict) -> dict:
    candidates = []
    for match in matches:
        for item in match.get("portfolio_candidates") or []:
            candidate = dict(item)
            candidate.setdefault("match_id", match.get("id"))
            candidates.append(candidate)

    open_bets = []
    for index, raw in enumerate((runtime.get("exposure") or {}).get("open_bets") or [], start=1):
        item = dict(raw)
        item["ticket_id"] = str(item.get("ticket_id") or item.get("id") or f"B{index:03d}")
        item["tier"] = normalize_layer(item.get("tier") or item.get("layer"), item.get("market"))
        item["status"] = "已锁单"
        item.setdefault("form", "单关")
        open_bets.append(item)

    layer_rows = []
    for name, meta in PORTFOLIO_LAYERS.items():
        proposed = [item for item in candidates if item.get("tier") == name]
        committed = [item for item in open_bets if item.get("tier") == name]
        layer_rows.append({
            "name": name, "label": name, **meta,
            "candidate_count": len(proposed),
            "locked_count": len(committed),
            "ticket_count": len(proposed) + len(committed),
            "candidate_exposure": sum(float(item.get("amount") or 0) for item in proposed),
            "locked_exposure": sum(float(item.get("amount") or 0) for item in committed),
            "status": "已有候选，仍未锁单" if proposed else ("已有锁单" if committed else "暂无合格候选"),
        })

    by_match: dict[str, list[dict]] = {}
    for item in candidates + open_bets:
        match_key = str(item.get("match_id") or item.get("match") or "").strip()
        if match_key:
            by_match.setdefault(match_key, []).append(item)
    overlap = []
    for match_key, items in by_match.items():
        if len(items) > 1:
            overlap.append({
                "match": items[0].get("match") or match_key,
                "ticket_ids": [item.get("ticket_id") for item in items],
                "risk": "同场多票",
                "control": "用完整赛果情景表检查共同失败风险与总暴露",
            })

    configured = runtime.get("betting_portfolio") or {}
    parlays = list(configured.get("parlays") or [])
    return {
        "layers": layer_rows,
        "candidates": candidates,
        "open_bets": open_bets,
        "parlays": parlays,
        "overlap_audit": overlap,
        "candidate_exposure": sum(float(item.get("amount") or 0) for item in candidates),
        "locked_exposure": float((runtime.get("exposure") or {}).get("current_open_exposure") or 0),
        "state": (
            "三层均为空仓" if not candidates and not open_bets and not parlays
            else "已有锁单，组合暴露生效" if open_bets
            else "组合候选待用户确认"
        ),
        "parlay_policy": "串关仅在每腿独立正EV、价格同时可得且相关性通过审核时成立",
    }


def find_report(schedule_match: dict, reports: dict[str, dict]) -> dict | None:
    exact = reports.get(f"{norm(schedule_match.get('homeTeam'))}|{norm(schedule_match.get('awayTeam'))}")
    if exact:
        return exact
    match_num = str(schedule_match.get("matchNum") or "")
    business_date = str(schedule_match.get("businessDate") or "")
    for report in reversed(list(reports.values())):
        match = report["payload"].get("match") or {}
        if match_num and match_num == str(match.get("match_num") or "") and business_date == str(match.get("business_date") or ""):
            return report
    return None


def find_report_for_pair(home: str, away: str, reports: dict[str, dict]) -> dict | None:
    exact = reports.get(f"{norm(home)}|{norm(away)}")
    if exact:
        return exact
    left, right = norm(home), norm(away)
    best: tuple[float, dict] | None = None
    for report in reports.values():
        match = report["payload"].get("match") or {}
        score = (
            SequenceMatcher(None, left, norm(match.get("home"))).ratio()
            + SequenceMatcher(None, right, norm(match.get("away"))).ratio()
        ) / 2
        if score >= 0.72 and (best is None or score > best[0]):
            best = (score, report)
    return best[1] if best else None


def completed_row(review: dict, report: dict | None, output_dir: Path, fallback_id: str) -> dict:
    label = str(review.get("赛事与对阵") or "")
    pair = re.split(r"\s+vs\s+", label.split("｜")[-1], maxsplit=1, flags=re.IGNORECASE)
    home, away = (pair + [""])[:2]
    return {
        "id": str(review.get("MatchID") or fallback_id),
        "home": home,
        "away": away,
        "result_90m": review.get("实际90分钟比分"),
        "after_extra_time": None,
        "kickoff": ((report or {}).get("payload") or {}).get("match", {}).get("kickoff_local"),
        "bet_locked": False,
        "classification": review.get("红黑与模型逻辑分类") or "模型复盘已记录",
        "prematch_report_url": relative_uri(report.get("html") if report else None, output_dir),
        "review": review,
    }


def relative_uri(path: Path | None, output_dir: Path) -> str:
    if not path:
        return ""
    # latest.html 固定在 data/match_workspace；稳定入口优先于历史快照内嵌浏览。
    return Path("..").joinpath(path.relative_to(DATA)).as_posix()


def pending_completed_row(home: str, away: str, kickoff: Any, report: dict | None,
                          output_dir: Path, row_id: str,
                          verified_results: dict[str, tuple[int, int]]) -> dict:
    score = verified_results.get(pair_key(home, away))
    return {
        "id": row_id,
        "home": home,
        "away": away,
        "result_90m": f"{score[0]}-{score[1]}" if score else None,
        "after_extra_time": None,
        "kickoff": kickoff,
        "bet_locked": False,
        "classification": "赛果已核验，复盘生成中" if score else "赛果待核验",
        "prematch_report_url": relative_uri(report.get("html") if report else None, output_dir),
        "review": None,
    }


def build(target_date: str, output_root: Path = OUTPUT) -> tuple[Path, Path]:
    runtime = load_json(RUNTIME, {})
    base_date = date.fromisoformat(target_date)
    schedule_sources = []
    schedule_refresh_times = []
    schedule_matches = []
    for offset in (0, 1):
        business_date = (base_date + timedelta(days=offset)).isoformat()
        source_path, source_payload = latest_schedule(business_date)
        if source_path:
            schedule_sources.append(source_path)
            if source_payload.get("fetch_time"):
                schedule_refresh_times.append(str(source_payload.get("fetch_time")))
            schedule_matches.extend(source_payload.get("matches") or [])
    unique_schedule_matches = {}
    for row in schedule_matches:
        identity = str(
            row.get("matchId")
            or "|".join(
                str(row.get(field) or "")
                for field in ("businessDate", "matchNum", "homeTeam", "awayTeam")
            )
        )
        unique_schedule_matches[identity] = row
    schedule = {
        "matches": sorted(
            unique_schedule_matches.values(),
            key=lambda row: (str(row.get("matchDate") or ""), str(row.get("matchTime") or "")),
        )
    }
    reports = latest_reports()
    reviews = review_rows(runtime)
    generated = datetime.now(SHANGHAI)
    stamp = generated.strftime("%Y%m%d_%H%M%S")
    output_dir = create_unique_output_dir(output_root, stamp)
    matches = []
    completed = []
    completed_ids: set[str] = set()
    verified_results = verified_result_map()
    schedule_keys = set()
    for row in schedule.get("matches") or []:
        home, away = row.get("homeTeam"), row.get("awayTeam")
        key = f"{norm(home)}|{norm(away)}"
        schedule_keys.add(key)
        report = find_report(row, reports)
        review = find_review(home, away, reviews)
        if review and review.get("实际90分钟比分"):
            item = completed_row(review, report, output_dir, f"schedule-{key}")
            completed.append(item)
            completed_ids.add(str(item["id"]))
            continue
        kickoff = f"{row.get('matchDate')} {str(row.get('matchTime') or '')[:5]}"
        if match_should_be_finished(kickoff, generated):
            item = pending_completed_row(home, away, kickoff, report, output_dir, f"schedule-{key}", verified_results)
            completed.append(item)
            completed_ids.add(str(item["id"]))
            continue
        summary = report_summary(report)
        matches.append({
            "id": str(row.get("matchId") or key), "match_num": row.get("matchNum"),
            "home": home, "away": away, "league": row.get("league"),
            "business_date": row.get("businessDate"),
            "kickoff": kickoff,
            "spf": row.get("spf") or {}, "rqspf": row.get("rqspf") or {},
            "official": True, "report_state": summary["state"], "primary": summary["primary"],
            "primary_error": summary["error"], "betting_state": summary["betting"],
            "report_url": relative_uri(report.get("html") if report else None, output_dir),
            "review": review,
            "portfolio_candidates": report_candidates(report, home, away),
        })
    # 已有分析但不在当天体彩列表中的关注场次仍可在同一工作台查看。
    for key, report in reports.items():
        match = report["payload"].get("match") or {}
        if match.get("business_date") not in {
            (base_date - timedelta(days=1)).isoformat(),
            target_date,
            (base_date + timedelta(days=1)).isoformat(),
        } or key in schedule_keys:
            continue
        if any(
            str(match.get("match_num") or "") == str(item.get("match_num") or "")
            and str(match.get("business_date") or "") == str(item.get("business_date") or "")
            for item in matches
        ):
            continue
        review = find_review(match.get("home"), match.get("away"), reviews)
        if review and review.get("实际90分钟比分"):
            item = completed_row(review, report, output_dir, f"report-{key}")
            if str(item["id"]) not in completed_ids:
                completed.append(item)
                completed_ids.add(str(item["id"]))
            continue
        kickoff = match.get("kickoff_local")
        if match_should_be_finished(kickoff, generated):
            item = pending_completed_row(
                match.get("home"), match.get("away"), kickoff, report, output_dir,
                f"report-{key}", verified_results,
            )
            if str(item["id"]) not in completed_ids:
                completed.append(item)
                completed_ids.add(str(item["id"]))
            continue
        summary = report_summary(report)
        matches.append({
            "id": str(match.get("live_match_id") or match.get("shuju_id") or key),
            "match_num": match.get("match_num") or "额外关注", "home": match.get("home"),
            "away": match.get("away"), "league": match.get("competition"),
            "business_date": match.get("business_date"), "kickoff": match.get("kickoff_local"),
            "spf": {}, "rqspf": {}, "official": False, "report_state": summary["state"],
            "primary": summary["primary"], "primary_error": summary["error"],
            "betting_state": summary["betting"],
            "report_url": relative_uri(report.get("html"), output_dir),
            "review": review,
            "portfolio_candidates": report_candidates(report, match.get("home"), match.get("away")),
        })
    for index, row in enumerate(runtime.get("latest_reviewed_matches") or [], start=1):
        name = str(row.get("match") or "")
        pair = re.split(r"\s+vs\s+", name, maxsplit=1, flags=re.IGNORECASE)
        home, away = (pair + [""])[:2]
        review = find_review(home, away, reviews)
        review_id = str((review or {}).get("MatchID") or f"review-{index}")
        if review_id in completed_ids:
            existing = next(item for item in completed if str(item.get("id")) == review_id)
            existing["after_extra_time"] = row.get("after_extra_time")
            existing["bet_locked"] = row.get("bet_locked") is True
            if not existing.get("kickoff"):
                existing_report = find_report_for_pair(home, away, reports)
                existing["kickoff"] = row.get("kickoff_local") or ((existing_report or {}).get("payload") or {}).get("match", {}).get("kickoff_local")
            continue
        report = find_report_for_pair(home, away, reports)
        completed.append({
            "id": review_id, "home": home, "away": away,
            "result_90m": row.get("result_90m"), "after_extra_time": row.get("after_extra_time"),
            "kickoff": row.get("kickoff_local") or ((report or {}).get("payload") or {}).get("match", {}).get("kickoff_local"),
            "bet_locked": row.get("bet_locked") is True,
            "classification": (
                row.get("review_classification")
                or (review or {}).get("红黑与模型逻辑分类")
                or "模型复盘已记录"
            ),
            "prematch_report_url": relative_uri(report.get("html") if report else None, output_dir),
            "review": review,
        })
        completed_ids.add(review_id)
    completed.sort(key=lambda item: str(item.get("kickoff") or ""), reverse=True)
    portfolio = build_daily_portfolio(matches, runtime)
    paper_root = DATA / "paper_ledger"
    paper_root.mkdir(parents=True, exist_ok=True)
    frozen_path = paper_root / "frozen.json"
    frozen_tickets = []
    if frozen_path.exists():
        try:
            frozen_tickets = (json.loads(frozen_path.read_text(encoding="utf-8")) or {}).get("tickets") or []
        except (OSError, json.JSONDecodeError, AttributeError):
            frozen_tickets = []
    sync_channel_price_overrides(frozen_tickets)
    price_overrides_payload = load_json(paper_root / "initial_price_overrides.json", {}) or {}
    initial_price_overrides = price_overrides_payload.get("tickets") or {}
    paper_ledger = build_paper_ledger(
        list(reports.values()),
        verified_result_map(),
        frozen_tickets=frozen_tickets,
        initial_price_overrides=initial_price_overrides,
    )
    frozen_records = paper_ledger.pop("_frozen_records", paper_ledger["tickets"])
    for item in matches:
        item.pop("portfolio_candidates", None)
    payload = {
        "model_name": runtime.get("model_name"), "model_version": runtime.get("model_version"),
        "balance": (runtime.get("bankroll") or {}).get("current_balance"),
        "open_bets": (runtime.get("exposure") or {}).get("open_bets") or [],
        "target_date": target_date, "generated_at": generated.isoformat(),
        "schedule_refreshed_at": max(schedule_refresh_times) if schedule_refresh_times else None,
        "published_as_latest": base_date >= date.today(),
        "automatic_analysis": False, "automatic_betting": False,
        "requires_explicit_lock_confirmation": True, "lock_state_changed": False,
        "schedule_source": [str(path.relative_to(ROOT)).replace("\\", "/") for path in schedule_sources],
        "available_cash": max(0.0, float((runtime.get("bankroll") or {}).get("current_balance") or 0) - portfolio["locked_exposure"]),
        "real_exposure": portfolio["locked_exposure"],
        "matches": matches, "completed": completed, "portfolio": portfolio,
        "paper_ledger": paper_ledger,
        "postmatch_dashboard_url": relative_uri(DATA / "postmatch_dashboard" / "latest.html", output_dir),
    }
    embedded = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    html_text = render(embedded)
    index = output_dir / "index.html"
    index.write_text(html_text, encoding="utf-8")
    (output_dir / "workspace.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (paper_root / "latest.json").write_text(json.dumps(paper_ledger, ensure_ascii=False, indent=2), encoding="utf-8")
    frozen_path.write_text(
        json.dumps({"schema_version": "1.0", "tickets": frozen_records}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    latest = output_root / "latest.html"
    publish_latest = base_date >= date.today()
    if publish_latest:
        shutil.copy2(index, latest)
        (output_root / "latest.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return index, latest


def render(embedded: str) -> str:
    return '''<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Football Betting OneShot｜盘口情报总览</title><style>
:root{--bg:#07111d;--panel:#0d1b2b;--panel2:#102238;--line:#263c56;--text:#edf7ff;--muted:#8ea4ba;--cyan:#39d7c2;--blue:#62a8ff;--gold:#f2c261;--good:#50d99a;--red:#ff7382}*{box-sizing:border-box}body{margin:0;background:radial-gradient(circle at 8% 0,#133c50 0,transparent 30%),radial-gradient(circle at 91% 3%,#1f2955 0,transparent 28%),var(--bg);color:var(--text);font:14px/1.55 "Microsoft YaHei",system-ui,sans-serif}.shell{max-width:1500px;margin:auto;padding:28px}.hero{display:flex;justify-content:space-between;gap:24px;align-items:flex-end;padding:10px 2px 24px}.eyebrow{color:var(--cyan);font-size:12px;font-weight:900;letter-spacing:.13em}.hero h1{font-size:36px;margin:5px 0}.hero p{margin:0;color:var(--muted)}.sync{min-width:290px;background:#0b1827;border:1px solid var(--line);border-radius:16px;padding:15px}.sync strong{display:block;color:var(--good);font-size:16px}.sync span{color:var(--muted)}.kpis{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-bottom:18px}.kpi{background:linear-gradient(145deg,var(--panel2),var(--panel));border:1px solid var(--line);border-radius:15px;padding:15px}.kpi b{display:block;font-size:24px;margin-bottom:2px}.kpi span{color:var(--muted);font-size:12px}.controls{position:sticky;top:0;z-index:5;display:flex;gap:9px;padding:11px 0;background:#07111ddd;backdrop-filter:blur(12px)}button,input{font:inherit}.controls button,.controls input,.action{border:1px solid var(--line);background:#0d1c2e;color:var(--text);border-radius:10px;padding:9px 12px}.controls button,.action{cursor:pointer}.controls button.active{border-color:var(--cyan);color:var(--cyan)}.controls input{flex:1}.card{background:linear-gradient(150deg,#0e1d30,#0a1726);border:1px solid var(--line);border-radius:18px;margin:14px 0;overflow:hidden}.card-title{display:flex;justify-content:space-between;align-items:center;padding:17px 19px;border-bottom:1px solid var(--line)}.card-title h2{font-size:18px;margin:0}.card-title span{color:var(--muted)}.table-wrap{overflow:auto}table{width:100%;border-collapse:collapse;min-width:960px}th{color:var(--muted);font-size:11px;text-align:left;padding:11px 15px;border-bottom:1px solid var(--line);letter-spacing:.04em}td{padding:14px 15px;border-bottom:1px solid #1e3047;vertical-align:middle}tbody tr:hover{background:#102239}tbody tr:last-child td{border-bottom:0}.match-name{font-weight:800;font-size:15px}.muted{color:var(--muted);font-size:12px}.badge{display:inline-flex;padding:3px 8px;border-radius:999px;font-size:11px;background:#24374d;color:#c7d8e9}.badge.good{background:#164637;color:#79efbd}.badge.gold{background:#473719;color:#ffd674}.badge.blue{background:#183d67;color:#93c8ff}.odds-inline{display:flex;gap:5px}.odd{min-width:46px;text-align:center;background:#091522;border:1px solid #1d324b;border-radius:7px;padding:4px 6px}.actions{display:flex;gap:7px;justify-content:flex-end}.action.primary{background:linear-gradient(105deg,#14877d,#2f67a7);border:0}.action.selected{background:#164637;color:#79efbd}.empty-row{text-align:center;color:var(--muted);padding:32px}.rules{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin:14px 0}.rule{background:#0a1726;border:1px solid var(--line);border-radius:14px;padding:13px}.rule b{display:block;margin-bottom:3px}.rule span{color:var(--muted);font-size:12px}footer{color:var(--muted);border-top:1px solid var(--line);padding:18px 2px;margin-top:18px}dialog{width:min(1380px,96vw);height:92vh;padding:0;border:1px solid var(--line);border-radius:18px;background:#091522;color:var(--text);box-shadow:0 30px 100px #000a}dialog::backdrop{background:#02060db8;backdrop-filter:blur(7px)}.dialog-head{display:flex;justify-content:space-between;align-items:center;padding:13px 16px;border-bottom:1px solid var(--line)}.dialog-head h3{margin:0}.dialog-tabs{display:flex;gap:7px}.dialog-tabs button,.close{border:1px solid var(--line);background:#11233a;color:var(--text);border-radius:9px;padding:8px 11px;cursor:pointer}.dialog-tabs button.active{color:var(--cyan);border-color:var(--cyan)}.dialog-body{height:calc(92vh - 65px);overflow:auto}.dialog-body iframe{width:100%;height:100%;border:0;background:#fff}.review{padding:28px}.review h2{margin:0 0 4px}.review-section{margin-top:18px}.review-section h3{margin:0 0 10px;color:var(--gold);font-size:15px}.review-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}.review-grid div{background:#0d1d30;border:1px solid var(--line);border-radius:12px;padding:13px}.review-grid label{display:block;color:var(--muted);font-size:11px}.review-grid p{margin:5px 0 0}.review-summary{background:#151225;border-left:3px solid var(--gold);border-radius:10px;padding:14px;margin-top:18px}.notice{color:var(--gold)}@media(max-width:900px){.shell{padding:15px}.hero{display:block}.sync{margin-top:15px}.kpis{grid-template-columns:repeat(2,1fr)}.rules{grid-template-columns:1fr}.hero h1{font-size:28px}.controls{flex-wrap:wrap}.controls input{min-width:100%}.dialog-tabs{overflow:auto}.dialog-head{align-items:flex-start}.review-grid{grid-template-columns:1fr}}
</style><style>
.portfolio-card{border-color:rgba(255,189,92,.28)}.portfolio-intro{display:flex;justify-content:space-between;gap:18px;align-items:center;padding:14px 19px;color:var(--muted);border-bottom:1px solid var(--line)}.portfolio-intro b{color:var(--text)}.portfolio-layers{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;padding:16px 19px}.portfolio-layer{position:relative;overflow:hidden;background:linear-gradient(145deg,#1a1529,#100d1b);border:1px solid var(--line);border-radius:15px;padding:16px}.portfolio-layer:before{content:"";position:absolute;inset:0 auto 0 0;width:3px;background:var(--layer-color)}.portfolio-layer .layer-head{display:flex;justify-content:space-between;align-items:center;gap:12px}.portfolio-layer h3{margin:0;font-size:16px}.portfolio-layer .layer-count{font-size:23px;font-weight:900;color:var(--layer-color)}.portfolio-layer p{margin:8px 0 0;color:var(--muted);font-size:12px}.portfolio-layer .layer-meta{display:flex;justify-content:space-between;gap:8px;margin-top:13px;padding-top:11px;border-top:1px solid var(--line);font-size:12px}.portfolio-sub{margin:0;padding:15px 19px 8px;font-size:14px}.portfolio-audits{display:grid;grid-template-columns:1fr 1fr;gap:12px;padding:0 19px 18px}.portfolio-audit{min-width:0;border:1px solid var(--line);border-radius:13px;overflow:hidden}.portfolio-audit h3{margin:0;padding:12px 14px;border-bottom:1px solid var(--line);font-size:14px}.portfolio-audit table{min-width:620px}.portfolio-audit td,.portfolio-audit th{padding:10px 12px}.portfolio-state{font-weight:800;color:var(--gold)}@media(max-width:900px){.portfolio-intro{display:block}.portfolio-intro span{display:block;margin-top:6px}.portfolio-layers,.portfolio-audits{grid-template-columns:1fr}}
</style></head><body><main class="shell"><header class="hero"><div><div class="eyebrow">FOOTBALL BETTING ONESHOT · INTELLIGENCE DASHBOARD</div><h1>盘口情报站 // 总览</h1><p id="subtitle"></p></div><div class="sync"><strong>● 体彩赛程每日更新</strong><span>只更新赛程 · 选中后才分析 · 未明确锁单不写注单</span></div></header><section class="kpis"><div class="kpi"><b id="upcomingCount">0</b><span>体彩已开售场次</span></div><div class="kpi"><b id="analyzedCount">0</b><span>已有赛前报告</span></div><div class="kpi"><b id="selectedCount">0</b><span>待分析清单</span></div><div class="kpi"><b id="completedCount">0</b><span>已完成复盘</span></div><div class="kpi"><b id="openBetCount">0</b><span>当前未结算注单</span></div><div class="kpi"><b id="balance">¥0</b><span>当前账户余额</span></div></section><section class="card portfolio-card"><div class="card-title"><h2>今日投注组合</h2><span class="portfolio-state" id="portfolioState">读取中</span></div><div class="portfolio-intro"><b>跨场资金组合，不属于任何单场报告</b><span>候选只按层展示；只有明确“锁单/已下单”才计入真实暴露与余额。</span></div><div class="portfolio-layers" id="portfolioLayers"></div><h3 class="portfolio-sub">候选与已锁票据</h3><div class="table-wrap"><table><thead><tr><th>票据</th><th>层级</th><th>比赛</th><th>玩法</th><th>赔率</th><th>EV</th><th>金额</th><th>形式</th><th>状态</th></tr></thead><tbody id="portfolioTickets"></tbody></table></div><div class="portfolio-audits"><div class="portfolio-audit"><h3>跨场串联审核</h3><div class="table-wrap"><table><thead><tr><th>票据</th><th>层级</th><th>串联</th><th>合成赔率</th><th>EV</th><th>相关性</th><th>状态</th></tr></thead><tbody id="portfolioParlays"></tbody></table></div></div><div class="portfolio-audit"><h3>同场重复暴露</h3><div class="table-wrap"><table><thead><tr><th>比赛</th><th>关联票据</th><th>风险</th><th>控制</th></tr></thead><tbody id="portfolioOverlap"></tbody></table></div></div></div></section><nav class="controls"><button class="filter active" data-filter="all">全部</button><button class="filter" data-filter="pending">未分析</button><button class="filter" data-filter="done">已分析</button><button class="filter" data-filter="selected">待分析</button><input id="search" placeholder="搜索球队、赛事、竞彩编号"></nav><section class="card"><div class="card-title"><h2>▶ 未开赛 · 体彩已开售（今日与下一业务日）</h2><span id="scheduleMeta"></span></div><div class="table-wrap"><table><thead><tr><th>开赛（北京）</th><th>比赛</th><th>体彩胜平负</th><th>状态</th><th>模型首推与错点</th><th>操作</th></tr></thead><tbody id="upcoming"></tbody></table></div></section><section class="card"><div class="card-title"><h2>✓ 已完赛 · 最近复盘</h2><button class="action" id="openAllReviews">打开完整复盘总表</button></div><div class="table-wrap"><table><thead><tr><th>比赛</th><th>90分钟比分</th><th>加时后</th><th>锁单状态</th><th>复盘结论</th><th>操作</th></tr></thead><tbody id="completed"></tbody></table></div></section><section class="rules"><div class="rule"><b>每天自动做什么</b><span>只刷新中国体彩竞彩足球赛程并重建总览页。</span></div><div class="rule"><b>什么时候分析</b><span>你点击“加入待分析”并明确要求后，才抓深层数据和生成报告。</span></div><div class="rule"><b>什么时候成为注单</b><span>只有你明确说“锁单/已下单”；候选、EV和网页选择都不算锁单。</span></div></section><footer>Football Betting OneShot · 纯数据分析与决策审计 · 页面不会自动下注或改变余额。</footer></main><dialog id="reportDialog"><div class="dialog-head"><div><h3 id="dialogTitle">比赛报告</h3><span class="muted" id="dialogMeta"></span></div><div class="dialog-tabs"><button id="tabPrematch">赛前分析</button><button id="tabReview">赛后复盘</button><button id="tabAllReviews">复盘总表</button><button class="close" id="closeDialog">关闭</button></div></div><div class="dialog-body" id="dialogBody"></div></dialog><script>
const DATA=''' + embedded + ''';let mode='all',current=null,selected=JSON.parse(localStorage.getItem('fbos-analysis-selections')||'[]');const $=s=>document.querySelector(s),key=m=>String(m.id),isSelected=m=>selected.some(x=>String(x.id)===key(m));$('#subtitle').textContent=`${DATA.model_version} · 体彩已开售 ${DATA.target_date} 起两个业务日 · 更新 ${new Date(DATA.generated_at).toLocaleString()}`;$('#scheduleMeta').textContent=`${DATA.target_date} 起 · ${DATA.matches.length} 场`;$('#balance').textContent=`¥${Number(DATA.balance||0).toFixed(2)}`;$('#openBetCount').textContent=DATA.open_bets.length;$('#completedCount').textContent=DATA.completed.length;
function analysisRequestUrl(m){const title=`[自动分析] ${m.home} vs ${m.away}`;const body=`match_id: ${m.id||''}\nbusiness_date: ${m.business_date||DATA.target_date}\nmatch: ${m.home} vs ${m.away}`;return `https://github.com/gemini077/football-betting-oneshot/issues/new?title=${encodeURIComponent(title)}&body=${encodeURIComponent(body)}`}
function saveSelection(m){if(m.report_state==='已分析'){if(m.report_url)openMatch(m,'prematch');return}if(!isSelected(m))selected.push({id:m.id,match_num:m.match_num,home:m.home,away:m.away,league:m.league,kickoff:m.kickoff,selected_at:new Date().toISOString(),analysis_requested:true});localStorage.setItem('fbos-analysis-selections',JSON.stringify(selected));fetch('http://127.0.0.1:8765/v1/analysis-selections',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({match:m})}).catch(()=>{});render();window.open(analysisRequestUrl(m),'_blank','noopener')}
function upcomingRow(m){const spf=m.spf||{},analyzed=m.report_state==='已分析';return `<tr data-row="${key(m)}"><td><b>${m.kickoff||'—'}</b><div class="muted">${m.match_num||'—'} · ${m.league||'—'}</div></td><td><div class="match-name">${m.home} <span class="muted">vs</span> ${m.away}</div><div class="muted">${m.official?'体彩在售':'额外关注'}</div></td><td><div class="odds-inline"><span class="odd">胜 ${spf.home??'—'}</span><span class="odd">平 ${spf.draw??'—'}</span><span class="odd">负 ${spf.away??'—'}</span></div></td><td><span class="badge ${analyzed?'good':'blue'}">${m.report_state}</span> ${!analyzed&&isSelected(m)?'<span class="badge gold">已提交</span>':''}</td><td><b>${m.primary||'—'}</b><div class="risk-line">错点：${m.primary_error||'—'}</div><div class="muted">${m.betting_state||'未锁单'}</div></td><td><div class="actions">${analyzed?`<button class="action primary" data-open="${key(m)}">打开报告</button>`:`<button class="action ${isSelected(m)?'selected':'primary'}" data-select="${key(m)}">${isSelected(m)?'重新提交分析':'分析这场'}</button><button class="action" data-open="${key(m)}">查看状态</button>`}</div></td></tr>`}
function completedRow(m){return `<tr><td><div class="match-name">${m.home} <span class="muted">vs</span> ${m.away}</div></td><td><b>${m.result_90m||'—'}</b></td><td>${m.after_extra_time||'—'}</td><td><span class="badge ${m.bet_locked?'gold':'blue'}">${m.bet_locked?'已锁单':'未锁单'}</span></td><td>${m.classification||'—'}</td><td><div class="actions"><button class="action primary" data-review="${m.id}">赛后复盘</button>${m.prematch_report_url?`<button class="action" data-prematch="${m.id}">赛前分析</button>`:''}</div></td></tr>`}
function numberText(v,d=2){const n=Number(v);return Number.isFinite(n)?n.toFixed(d):'—'}function pctText(v){const n=Number(v);return Number.isFinite(n)?`${(n*100).toFixed(1)}%`:'—'}function renderPortfolio(){const p=DATA.portfolio||{},layers=p.layers||[];$('#portfolioState').textContent=p.state||'三层均为空仓';$('#portfolioLayers').innerHTML=layers.map(layer=>`<article class="portfolio-layer" style="--layer-color:${layer.color||'#ffbd5c'}"><div class="layer-head"><h3>${layer.name}</h3><span class="layer-count">${layer.ticket_count||0}</span></div><p>${layer.role||''}</p><div class="layer-meta"><span>候选 ${layer.candidate_count||0}</span><span>已锁 ${layer.locked_count||0}</span><span>暴露 ¥${numberText(layer.locked_exposure||0)}</span></div></article>`).join('')||'<div class="empty-row">暂无层级配置</div>';const tickets=[...(p.candidates||[]),...(p.open_bets||[])];$('#portfolioTickets').innerHTML=tickets.map(t=>`<tr><td><b>${t.ticket_id||t.id||'—'}</b></td><td><span class="badge gold">${t.tier||'中轴层'}</span></td><td>${t.match||'—'}</td><td>${t.market||t.selection||'—'}</td><td>${numberText(t.user_channel_odds??t.observed_odds??t.odds)}</td><td>${pctText(t.repriced_ev??t.initial_ev??t.ev)}</td><td>${t.amount==null?'—':`¥${numberText(t.amount)}`}</td><td>${t.form||'单关'}</td><td><span class="badge ${String(t.status||'').includes('锁')?'good':'blue'}">${t.status||'候选未锁单'}</span></td></tr>`).join('')||'<tr><td colspan="9" class="empty-row">当前没有通过阈值的候选，也没有已锁票据</td></tr>';$('#portfolioParlays').innerHTML=(p.parlays||[]).map(t=>`<tr><td><b>${t.ticket_id||t.id||'—'}</b></td><td>${t.tier||'—'}</td><td>${(t.legs||[]).map(x=>typeof x==='string'?x:(x.label||x.market||x.match||'—')).join(' × ')||'—'}</td><td>${numberText(t.combined_odds)}</td><td>${pctText(t.ev)}</td><td>${t.correlation||'待审计'}</td><td>${t.status||'候选未锁单'}</td></tr>`).join('')||'<tr><td colspan="7" class="empty-row">没有通过审核的跨场串联；不为凑票强行组合</td></tr>';$('#portfolioOverlap').innerHTML=(p.overlap_audit||[]).map(t=>`<tr><td>${t.match||'—'}</td><td>${(t.ticket_ids||[]).join('、')||'—'}</td><td>${t.risk||'—'}</td><td>${t.control||'—'}</td></tr>`).join('')||'<tr><td colspan="4" class="empty-row">当前没有同场重复暴露</td></tr>'}
function render(){const q=$('#search').value.trim().toLowerCase();const rows=DATA.matches.filter(m=>(!q||JSON.stringify(m).toLowerCase().includes(q))&&(mode==='all'||mode==='pending'&&m.report_state!=='已分析'||mode==='done'&&m.report_state==='已分析'||mode==='selected'&&isSelected(m)));$('#upcoming').innerHTML=rows.map(upcomingRow).join('')||'<tr><td colspan="6" class="empty-row">没有符合条件的场次</td></tr>';$('#completed').innerHTML=DATA.completed.map(completedRow).join('')||'<tr><td colspan="6" class="empty-row">暂无已完成复盘</td></tr>';$('#upcomingCount').textContent=DATA.matches.length;$('#analyzedCount').textContent=DATA.matches.filter(m=>m.report_state==='已分析').length;$('#selectedCount').textContent=selected.length;document.querySelectorAll('[data-select]').forEach(b=>b.onclick=()=>saveSelection(DATA.matches.find(m=>key(m)===b.dataset.select)));document.querySelectorAll('[data-open]').forEach(b=>b.onclick=()=>openMatch(DATA.matches.find(m=>key(m)===b.dataset.open),'prematch'));document.querySelectorAll('[data-review]').forEach(b=>b.onclick=()=>openCompleted(DATA.completed.find(x=>String(x.id)===b.dataset.review),'review'));document.querySelectorAll('[data-prematch]').forEach(b=>b.onclick=()=>openCompleted(DATA.completed.find(x=>String(x.id)===b.dataset.prematch),'prematch'))}
function openReport(url){if(!url)return false;window.location.assign(url);return true}function showEmpty(text){$('#dialogBody').innerHTML=`<div class="empty-row" style="padding:80px">${text}</div>`}function field(label,value){return `<div><label>${label}</label><p>${value??'—'}</p></div>`}function showPrematch(){setTab('tabPrematch');const url=current?.prematch_report_url||current?.report_url;if(!url)return showEmpty('该场没有可用的赛前分析报告。');openReport(url)}function showReview(){setTab('tabReview');const r=current?.review;if(!r)return showEmpty('该场尚未生成赛后复盘；比赛结束后的单次任务会更新这里。');const t=r._timeline||{},root=r._root_cause||{};$('#dialogBody').innerHTML=`<div class="review"><h2>${r['赛事与对阵']||current.home+' vs '+current.away}</h2><div class="muted">${r.MatchID||current.id||'—'} · 严格按赛前唯一合约与90分钟赛果结算</div><section class="review-section"><h3>赛果与严格结算审计</h3><div class="review-grid">${field('实际90分钟比分',r['实际90分钟比分'])}${field('主维度严格结算',r['主维度是否命中'])}${field('唯一比分精确命中',r['比分是否命中'])}${field('相邻比分误差（仍未中）',r['相邻比分污染'])}${field('冷门/右尾诊断（不替代命中）',r['冷门/右尾污染'])}${field('模型逻辑分类（独立于结算）',r['红黑与模型逻辑分类'])}</div></section><section class="review-section"><h3>赛前各维度逐项复核</h3><div class="review-grid">${field('三维交叉验证',r['赛前三维交叉验证结论'])}${field('盘路与价格纪律',r['盘路性质判定'])}${field('唯一比分',r['赛前唯一首推比分'])}${field('唯一主维度',r['赛前首推主维度'])}${field('亚盘方向（独立复核）',r['赛前亚盘方向'])}${field('大小球（独立复核）',r['赛前大小球方向'])}${field('双方进球（独立复核）',r['赛前BTTS判断'])}${field('赛前最大错点',r['赛前最大错点'])}${field('赛后错点归因',r['错点归因（单选）'])}</div></section><section class="review-section"><h3>盘口时间轴与数据有效性</h3><div class="review-grid">${field('开赛倒计时',t['开赛倒计时'])}${field('锁单窗口',t['锁单窗口合规性'])}${field('初盘定位',t['初盘定位'])}${field('终盘定位',t['终盘定位（临场15min）'])}${field('终盘相对初盘',t['终盘对比初盘变化'])}${field('经验库触发',t['经验库触发类型'])}${field('盘口变化与赛果',t['盘口变化与赛果方向'])}${field('欧亚理论校验',t['欧亚理论盘型校验'])}${field('最终赛果验证',t['最终赛果验证'])}${field('数据完整度',t['数据完整度'])}${field('来源与备注',t['来源/备注'])}</div></section><section class="review-section"><h3>根因、反事实与模型修正</h3><div class="review-grid">${field('决策节点审计',root['决策节点审计'])}${field('反事实推演',root['反事实推演'])}${field('赛前可识别性',root['赛前可识别性'])}${field('是否修改模型',root['是否修改模型'])}${field('具体修改建议',root['具体修改建议'])}${field('收敛结论',root['收敛结论'])}${field('生效状态',root['生效状态'])}${field('修正优先级',root['优先级'])}${field('最大错点类型',r['最大错点类型'])}${field('错点触发透视',root['最大错点触发透视'])}</div></section><div class="review-summary"><b>完整复盘摘要</b><p>${r['复盘摘要']||'—'}</p></div><p class="notice">未锁单场次只复盘模型，不制造注单盈亏；辅助玩法命中不替代唯一主维度结算。</p></div>`}function openAllReviews(){if(!current)current=DATA.completed.find(x=>x.review)||null;$('#dialogTitle').textContent='赛后复盘总表';$('#dialogMeta').textContent='记录ID排序 · 90分钟口径';setTab('tabAllReviews');$('#dialogBody').innerHTML=`<iframe title="赛后复盘总表" src="${DATA.postmatch_dashboard_url}"></iframe>`;$('#reportDialog').showModal()}function openMatch(m,view){current=m;$('#dialogTitle').textContent=`${m.home} vs ${m.away}`;$('#dialogMeta').textContent=`${m.match_num||'—'} · ${m.league||'—'} · ${m.kickoff||'—'}`;$('#reportDialog').showModal();view==='review'?showReview():showPrematch()}function openCompleted(m,view='review'){current=m;$('#dialogTitle').textContent=`${m.home} vs ${m.away}`;$('#dialogMeta').textContent=`90分钟 ${m.result_90m||'—'}`;$('#reportDialog').showModal();view==='prematch'?showPrematch():showReview()}function setTab(id){['tabPrematch','tabReview','tabAllReviews'].forEach(x=>$('#'+x).classList.toggle('active',x===id))}
$('#tabPrematch').onclick=()=>current?showPrematch():showEmpty('请先选择比赛');$('#tabReview').onclick=()=>current?showReview():showEmpty('请先选择比赛');$('#tabAllReviews').onclick=openAllReviews;$('#openAllReviews').onclick=openAllReviews;$('#closeDialog').onclick=()=>$('#reportDialog').close();$('#search').oninput=render;document.querySelectorAll('.filter').forEach(b=>b.onclick=()=>{mode=b.dataset.filter;document.querySelectorAll('.filter').forEach(x=>x.classList.toggle('active',x===b));render()});renderPortfolio();render();
</script><style>
:root{--bg:#0a0811;--panel:#13101f;--panel2:#181326;--line:rgba(255,255,255,.10);--text:#f7f2fa;--muted:#8c839f;--cyan:#ff7189;--blue:#9b7fd0;--gold:#ffbd5c;--good:#39d6a0;--red:#ff3657}body{background:radial-gradient(70% 42% at 50% -8%,rgba(255,54,87,.26),transparent 68%),linear-gradient(180deg,#171027 0,#0a0811 52%)}.sync,.rule{background:#13101f}.kpi{background:linear-gradient(145deg,#181326,#13101f)}.controls{background:#0a0811dd}.controls button,.controls input,.action{background:#181326}.card{background:linear-gradient(180deg,rgba(255,255,255,.035),rgba(255,255,255,.018))}td{border-bottom-color:rgba(255,255,255,.07)}tbody tr:hover{background:rgba(255,54,87,.045)}.action.primary{background:linear-gradient(105deg,#ff3657,#9b7fd0)}.badge.good{background:rgba(57,214,160,.14);color:#6ce7bb}.badge.blue{background:rgba(155,127,208,.15);color:#c8b8e7}.odd{background:#100b1b;border-color:rgba(255,255,255,.11)}dialog{background:#0a0811}.dialog-tabs button,.close{background:#181326}.review-grid div{background:#13101f}.risk-line{color:var(--gold);font-size:11px;line-height:1.45;margin:5px 0;max-width:420px}
</style><style>
.hero{position:relative;min-height:310px;padding:42px;overflow:hidden;border:1px solid rgba(255,54,87,.22);border-radius:18px;align-items:flex-end;background:radial-gradient(55% 80% at 50% 0,rgba(255,54,87,.28),transparent 70%),linear-gradient(180deg,#21132f,#0d0915)}.hero:after{content:"";position:absolute;left:-3%;right:-3%;bottom:-1px;height:45%;background:linear-gradient(145deg,#2b1a3a,#120c1c);clip-path:polygon(0 72%,10% 36%,23% 69%,34% 28%,47% 73%,61% 31%,74% 70%,87% 25%,100% 61%,100% 100%,0 100%);opacity:.88}.hero>div{position:relative;z-index:1}.hero h1{font-size:40px;letter-spacing:.035em;text-shadow:0 0 28px rgba(255,54,87,.35)}.hero h1 span{color:var(--red);font-weight:500}.hero-chips{display:flex;flex-wrap:wrap;gap:9px;margin-top:18px}.hero-chips span{padding:6px 11px;border:1px solid rgba(255,255,255,.14);border-radius:6px;background:rgba(255,255,255,.055);font-size:12px;color:var(--muted)}.hero-chips b{color:var(--text)}.account-kpis{margin-top:20px}.paper-card{border-color:rgba(255,54,87,.3)}.paper-card .card-title{background:linear-gradient(90deg,rgba(255,54,87,.09),transparent)}.section-kicker{font-size:10px;letter-spacing:.24em;color:var(--cyan);font-weight:800}.paper-separation{color:var(--good)!important;border:1px solid rgba(57,214,160,.32);border-radius:999px;padding:5px 10px}.paper-intro{display:flex;justify-content:space-between;gap:16px;padding:14px 19px;border-bottom:1px solid var(--line);color:var(--muted)}.paper-intro b{color:var(--text)}.paper-kpis{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;padding:17px 19px 8px}.paper-kpi{border:1px solid var(--line);border-radius:12px;padding:14px;background:rgba(255,255,255,.025)}.paper-kpi b{display:block;font-size:23px;color:var(--text)}.paper-kpi b.positive{color:var(--good)}.paper-kpi b.negative{color:var(--red)}.paper-kpi span{font-size:11px;color:var(--muted)}.paper-grid{display:grid;grid-template-columns:1fr;gap:14px;padding:4px 19px 14px}.paper-grid>div{min-width:0;border:1px solid var(--line);border-radius:13px;overflow:hidden;background:rgba(0,0,0,.08)}.paper-subhead{display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid var(--line)}.paper-grid .portfolio-sub{padding:12px 14px;margin:0;border-bottom:0}.paper-more{margin-right:12px;border:1px solid rgba(255,255,255,.15);border-radius:8px;padding:6px 11px;background:#21182f;color:var(--text);cursor:pointer}.paper-more:hover{border-color:var(--cyan);color:var(--cyan)}.compact-table{min-width:620px}.compact-table th,.compact-table td{padding:10px 12px}.paper-note{margin:0;padding:0 19px 17px;color:var(--muted);font-size:11px}.result-win{color:var(--good);font-weight:800}.result-loss{color:var(--red);font-weight:800}.result-pending{color:var(--gold);font-weight:800}.result-observe{color:var(--blue);font-weight:800}#paperTicketsDialog{width:min(1050px,94vw)}#paperTicketsDialog .table-wrap{max-height:68vh}.paper-dialog-note{padding:0 18px 14px;color:var(--muted);font-size:11px}
@media(max-width:1000px){.paper-kpis{grid-template-columns:repeat(3,1fr)}.paper-grid{grid-template-columns:1fr}}@media(max-width:700px){.hero{min-height:270px;padding:28px 20px}.hero h1{font-size:29px}.paper-intro{display:block}.paper-intro span{display:block;margin-top:6px}.paper-kpis{grid-template-columns:repeat(2,1fr)}}
</style><script>
(function(){
  const hero=document.querySelector('.hero');
  hero.querySelector('.eyebrow').textContent='FOOTBALL BETTING ONESHOT · MODEL LEDGER';
  hero.querySelector('h1').innerHTML='赛前分析 <span>//</span> 模拟验证 <span>//</span> 严格复盘';
  hero.querySelector('.sync strong').textContent='● 赛程与赛果自动更新';
  hero.querySelector('.sync span').textContent='选中后才分析 · 模拟账不影响余额 · 未明确锁单不写真实注单';
  const heroChips=document.createElement('div');
  heroChips.className='hero-chips';
  const ps=(DATA.paper_ledger||{}).summary||{};
  heroChips.innerHTML=`<span>模型 <b>${DATA.model_version||'—'}</b></span><span>业务日 <b>${DATA.target_date||'—'}</b></span><span>待验赛 <b>${ps.pending||0}</b></span>`;
  hero.querySelector('h1').insertAdjacentElement('afterend',heroChips);
  const scheduleTime=DATA.schedule_refreshed_at?new Date(DATA.schedule_refreshed_at).toLocaleString():'沿用最近成功赛程';
  $('#subtitle').textContent=`模拟账与真实账分离 · 已结模拟 ${ps.settled||0} 注 · 赛程数据 ${scheduleTime} · 页面生成 ${new Date(DATA.generated_at).toLocaleString()}`;
  const account=document.querySelector('.kpis');
  account.classList.add('account-kpis');
  account.innerHTML=`<div class="kpi"><b id="balance">¥${Number(DATA.balance||0).toFixed(2)}</b><span>当前账户余额</span></div><div class="kpi"><b>¥${Number(DATA.available_cash||0).toFixed(2)}</b><span>可用现金</span></div><div class="kpi"><b>¥${Number(DATA.real_exposure||0).toFixed(2)}</b><span>真实锁单暴露</span></div><div class="kpi"><b id="openBetCount">${(DATA.open_bets||[]).length}</b><span>真实未结注单</span></div><div class="kpi"><b id="analyzedCount">${DATA.matches.filter(m=>m.report_state==='已分析').length}</b><span>已有赛前报告</span></div><div class="kpi"><b id="completedCount">${DATA.completed.length}</b><span>已完成复盘</span></div><span id="upcomingCount" hidden>${DATA.matches.length}</span><span id="selectedCount" hidden>${selected.length}</span>`;
  const paper=document.createElement('section');
  paper.className='card paper-card';
  paper.innerHTML=`<div class="card-title"><div><div class="section-kicker">PAPER PERFORMANCE</div><h2>模型模拟战绩</h2></div><span class="paper-separation">与真实账户完全分离</span></div><div class="paper-intro"><b>报告冻结即登记，赛后不可改方向</b><span>先冻结赛前赔率并计算EV，再按1/4凯利决定金额；最低2.00元、步进0.01元，负EV或不足最低额直接放弃。</span></div><div class="paper-kpis" id="paperKpis"></div><div class="paper-grid"><div><div class="paper-subhead"><h3 class="portfolio-sub">分玩法表现</h3></div><div class="table-wrap"><table class="compact-table"><thead><tr><th>玩法</th><th>已结</th><th>胜-负</th><th>命中率</th><th>ROI</th><th>盈亏</th></tr></thead><tbody id="paperGroups"></tbody></table></div></div><div><div class="paper-subhead"><h3 class="portfolio-sub">最近模拟注单</h3><button type="button" class="paper-more" id="openPaperTickets">查看全部</button></div><div class="table-wrap"><table class="compact-table"><thead><tr><th>比赛</th><th>冻结合约</th><th>赔率</th><th>EV</th><th>金额</th><th>结算</th><th>盈亏</th></tr></thead><tbody id="paperTickets"></tbody></table></div></div></div><p class="paper-note">模拟战绩用于检验模型与价格纪律，不代表真实投注收益；命中率不含走盘，最终价值以ROI、回撤和样本量共同判断。</p>`;
  const portfolio=document.querySelector('.portfolio-card');
  portfolio.parentNode.insertBefore(paper,portfolio);
  const signed=v=>{const n=Number(v);return Number.isFinite(n)?`${n>0?'+':''}${n.toFixed(2)}`:'—'};
  const rate=v=>Number.isFinite(Number(v))?`${(Number(v)*100).toFixed(1)}%`:'—';
  const tone=v=>Number(v)>0?'positive':Number(v)<0?'negative':'';
  $('#paperKpis').innerHTML=[['待结算',ps.pending||0,''],['已结算',ps.settled||0,''],['模拟盈亏',signed(ps.profit_units)+' 元',tone(ps.profit_units)],['模拟ROI',rate(ps.roi),tone(ps.roi)],['命中率',rate(ps.hit_rate),''],['最大回撤',Number(ps.max_drawdown_units||0).toFixed(2)+' 元','negative']].map(x=>`<div class="paper-kpi"><b class="${x[2]}">${x[1]}</b><span>${x[0]}</span></div>`).join('');
  $('#paperGroups').innerHTML=((DATA.paper_ledger||{}).groups||[]).map(g=>`<tr><td><b>${g.market_group}</b></td><td>${g.settled}</td><td>${g.wins}-${g.losses}</td><td>${rate(g.hit_rate)}</td><td class="${tone(g.roi)}">${rate(g.roi)}</td><td class="${tone(g.profit_units)}">${signed(g.profit_units)} 元</td></tr>`).join('')||'<tr><td colspan="6" class="empty-row">尚无可结算模拟样本</td></tr>';
  const statusMeta=t=>t.status==='settled'?[t.settlement,(Number(t.profit_units)>0?'result-win':Number(t.profit_units)<0?'result-loss':'result-pending')]:t.status==='pending'?['待赛果','result-pending']:t.status==='rejected_by_ev'?[t.settlement||'EV不通过','result-loss']:['观察无价','result-observe'];
  const allPaperTickets=((DATA.paper_ledger||{}).tickets||[]).slice().reverse();
  const ticketRows=rows=>rows.map(t=>{const sm=statusMeta(t);const ev=t.sizing_ev==null?t.ev:t.sizing_ev;return `<tr><td><b>${t.match}</b><div class="muted">${t.model_version||'—'} · ${t.ticket_id}</div></td><td>${t.market}<div class="muted">${t.price_source||'无有效赛前价格'}</div></td><td>${t.odds==null?'—':Number(t.odds).toFixed(2)}</td><td class="${tone(ev)}">${rate(ev)}</td><td>${Number(t.stake_units||0).toFixed(2)}</td><td class="${sm[1]}">${sm[0]}</td><td class="${tone(t.profit_units)}">${t.profit_units==null?'—':signed(t.profit_units)+' 元'}</td></tr>`;}).join('')||'<tr><td colspan="7" class="empty-row">生成赛前报告后，模拟注单会自动出现在这里</td></tr>';
  $('#paperTickets').innerHTML=ticketRows(allPaperTickets.slice(0,4));
  const paperDialog=document.createElement('dialog');
  paperDialog.id='paperTicketsDialog';
  paperDialog.innerHTML=`<div class="dialog-head"><div><b>全部模拟注单</b><div class="muted">共 ${allPaperTickets.length} 条 · 方向、赔率、EV与金额按赛前冻结记录展示</div></div><button type="button" class="close" id="closePaperTickets">关闭</button></div><div class="table-wrap"><table class="compact-table"><thead><tr><th>比赛</th><th>冻结合约</th><th>赔率</th><th>EV</th><th>金额</th><th>结算</th><th>盈亏</th></tr></thead><tbody>${ticketRows(allPaperTickets)}</tbody></table></div><p class="paper-dialog-note">模拟金额采用100元独立纸面本金、1/4凯利、单注上限5%；最低2.00元、步进0.01元，负EV与低于最低额均不投注。</p>`;
  document.body.appendChild(paperDialog);
  $('#openPaperTickets').onclick=()=>paperDialog.showModal();
  $('#closePaperTickets').onclick=()=>paperDialog.close();
  portfolio.querySelector('.portfolio-intro b').textContent='保本／中轴／博上属于真实决策层';
  portfolio.querySelector('.portfolio-intro span').textContent='模拟注单不会自动进入这里；只有明确“锁单/已下单”才计入真实暴露与余额。';
  // Homepage interaction cleanup: analyzed fixtures have one report action,
  // completed fixtures have one unified report action, and kickoff is visible.
  const upcomingSection=document.querySelector('#upcoming').closest('section.card');
  const completedSection=document.querySelector('#completed').closest('section.card');
  upcomingSection.querySelector('.card-title h2').textContent='\u25b6 \u672a\u5f00\u8d5b';
  completedSection.querySelector('.card-title h2').textContent='\u2713 \u5df2\u5b8c\u8d5b';
  completedSection.querySelector('thead tr').innerHTML='<th>\u5f00\u8d5b\uff08\u5317\u4eac\uff09</th><th>\u6bd4\u8d5b</th><th>90\u5206\u949f\u6bd4\u5206</th><th>\u52a0\u65f6\u540e</th><th>\u9501\u5355\u72b6\u6001</th><th>\u590d\u76d8\u7ed3\u8bba</th><th>\u64cd\u4f5c</th>';
  document.querySelector('.rules')?.remove();
  upcomingRow=m=>{const spf=m.spf||{},analyzed=m.report_state==='已分析',hasReport=Boolean(m.report_url);const actions=analyzed?`<button class="action primary" data-open="${key(m)}">\u6253\u5f00\u62a5\u544a</button>`:`<button class="action ${isSelected(m)?'selected':'primary'}" data-select="${key(m)}">${isSelected(m)?'\u91cd\u65b0\u63d0\u4ea4\u5206\u6790':'\u52a0\u5165\u5f85\u5206\u6790'}</button>${hasReport?`<button class="action" data-open="${key(m)}">\u67e5\u770b\u6570\u636e\u72b6\u6001</button>`:''}`;return `<tr data-row="${key(m)}"><td><b>${m.kickoff||'\u2014'}</b><div class="muted">${m.match_num||'\u2014'} \u00b7 ${m.league||'\u2014'}</div></td><td><div class="match-name">${m.home} <span class="muted">vs</span> ${m.away}</div><div class="muted">${m.official?'\u4f53\u5f69\u5728\u552e':'\u989d\u5916\u5173\u6ce8'}</div></td><td><div class="odds-inline"><span class="odd">\u80dc ${spf.home??'\u2014'}</span><span class="odd">\u5e73 ${spf.draw??'\u2014'}</span><span class="odd">\u8d1f ${spf.away??'\u2014'}</span></div></td><td><span class="badge ${analyzed?'good':'blue'}">${m.report_state}</span> ${isSelected(m)?'<span class="badge gold">\u5df2\u9009\u62e9</span>':''}</td><td><b>${m.primary||'\u2014'}</b><div class="risk-line">\u9519\u70b9\uff1a${m.primary_error||'\u2014'}</div><div class="muted">${m.betting_state||'\u672a\u9501\u5355'}</div></td><td><div class="actions">${actions}</div></td></tr>`};
  completedRow=m=>{const action=m.review?`<button class="action primary" data-review="${m.id}">\u6253\u5f00\u62a5\u544a</button>`:(m.prematch_report_url?`<button class="action" data-prematch="${m.id}">\u6253\u5f00\u62a5\u544a</button>`:'<span class="muted">\u5f85\u590d\u76d8</span>');return `<tr><td><b>${m.kickoff||'\u2014'}</b></td><td><div class="match-name">${m.home} <span class="muted">vs</span> ${m.away}</div></td><td><b>${m.result_90m||'\u5f85\u6838\u9a8c'}</b></td><td>${m.after_extra_time||'\u2014'}</td><td><span class="badge ${m.bet_locked?'gold':'blue'}">${m.bet_locked?'\u5df2\u9501\u5355':'\u672a\u9501\u5355'}</span></td><td>${m.classification||'\u2014'}</td><td>${action}</td></tr>`};
  render();
})();
</script></body></html>'''


def main() -> int:
    parser = argparse.ArgumentParser(description="生成统一赛程、赛前分析和赛后复盘工作台")
    parser.add_argument("--date", default=date.today().isoformat())
    parser.add_argument("--output-root", default=str(OUTPUT))
    args = parser.parse_args()
    index, latest = build(args.date, Path(args.output_root))
    print(json.dumps({
        "index": str(index), "latest": str(latest),
        "published_as_latest": date.fromisoformat(args.date) >= date.today(),
        "automatic_analysis": False, "lock_state_changed": False,
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
